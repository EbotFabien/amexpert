import sys
import datetime
import openpyxl
import xlrd,xlwt
from project.data_base_ import db
from project.data_base_.Models import Tarifs,Mission,Client,Expert,Client_History,prospect,prospect_History,Expert_History,Tarif_base
import flask as pd
from flask import Flask,render_template,url_for,flash,redirect,request,Blueprint

def regex1(data,Type):
    if  Type=='str':
        if type(data)==str:
            if data.isalnum() == False:
                a=str(data).split()
                vii=''.join(a)
                vii=vii.split("-")
                vii=''.join(vii)
                return vii.lower()
            if data == 'None':
                return ''
            if data == 'xxx':
                return ''
            if data == 'XXX':
                return ''
            else:
                return data.lower()
        else:
            return ''
    
    if  Type=='str1':
        if type(data)==str:
            if data == 'xxx':
                return ''
            if data == 'XXX':
                return ''
            if data == 'None':
                return ''
            else:
                return data.lower()
        else:
            return ''

    if  Type=='int':
        if type(data) == int:
            return int(data)
        if type(data) == str:
            if data.isnumeric() == True:
                return int(data)
            a="".join(data.split())
            if a.isnumeric() == True:
                return int(a) 
            else:
                return 'Failed'
        if data == None:
            return None 
        else:
            return 'Failed'
        
    if  Type=='M':
        if data == '':
            return None
        if type(data) == int:
            return int(data)
        if type(data) == float:
            return int(data)
        if type(data) == str:
            if data.isnumeric() == True:
                return int(data)
            a="".join(data.split())
            if a.isnumeric() == True:
                return int(a) 
            else:
                return 0
        if data == None:
            return None 
        else:
            return 0            
            
    
    if  Type=='dec':
        
        if isinstance(data,int) == True:
            return round(data,2)
        if isinstance(data,float) == True:
            return data
        if data == '':
            return ''
        if data == None:
            return None
        else:
            return 'Failed'
    
    if  Type=='S':
        if isinstance(data,int) == True:
            return round(data,2)
        if isinstance(data,float) == True:
            return data
        if data == '':
            return None
        if type(data)==str:
            return None
        if data == None:
            return None
        else:
            return None

    if  Type == 'perc':
        if type(data) == str:
            if data == '':
                return 0.00
            if data[-1] == '%':
                return float(data.strip('%'))/100
            else:
                return 0.00
        if type(data) == int:
            return data/100
        if type(data) == float:
            if data % 1 !=0:
                return data
            else:
                return data/100
        else:
                return 0.00
    if  Type=='date':
        
        if isinstance(data,datetime.datetime) == True:
            return data
        elif isinstance(data,datetime.date) == True:
            return data
        elif data == None:
            a=1
        else:
            
            return False
            

def failed(av):
    ba=[]
    v=0
    for oo in av:
        tr=len(oo)
    for i in range(0,tr):
        ba.append(i)
    wb = xlwt.Workbook()
    style1 = xlwt.easyxf(num_format_str='D-MMM-YY')
    ws = wb.add_sheet('Anomalie')
    for oo in av:
            for q,i in zip(ba,oo) :
                if isinstance(i,datetime.datetime) == True:
                    ws.write(v, q, i,style1)
                elif isinstance(i,datetime.date) == True:
                    ws.write(v, q, i,style1) 
                else:
                    ws.write(v, q, i)
            v=v+1
    #wb.save('C:/Users/user/Downloads/Telegram Desktop/Anomalieclient.xls')

def failed1(av):
    ba=[]
    v=0
    for oo in av:
        tr=len(oo)
    for i in range(0,tr):
        ba.append(i)
    wb = xlwt.Workbook()
    style1 = xlwt.easyxf(num_format_str='D-MMM-YY')
    ws = wb.add_sheet('Anomalie')
    for oo in av:
            for q,i in zip(ba,oo) :
                if isinstance(i,datetime.datetime) == True:
                    ws.write(v, q, i,style1)
                elif isinstance(i,datetime.date) == True:
                    ws.write(v, q, i,style1) 
                else:
                    ws.write(v, q, i)
            v=v+1
    #wb.save('C:/Users/user/Downloads/Telegram Desktop/AnomaliePROSPECT.xls')

def good1(av):
    ba=[]
    v=0
    for oo in av:
        tr=len(oo)
    for i in range(0,tr):
        ba.append(i)
    wb = xlwt.Workbook()
    style1 = xlwt.easyxf(num_format_str='D-MMM-YY')
    ws = wb.add_sheet('Bien')
    for oo in av:
            for q,i in zip(ba,oo) :
                if isinstance(i,datetime.datetime) == True:
                    ws.write(v, q, i,style1)
                elif isinstance(i,datetime.date) == True:
                    ws.write(v, q, i,style1) 
                else:
                    ws.write(v, q, i)
            v=v+1
    #wb.save('C:/Users/user/Downloads/Telegram Desktop/BienPROSPECT.xls')

def good2(av):
    ba=[]
    v=0
    for oo in av:
        tr=len(oo)
    for i in range(0,tr):
        ba.append(i)
    wb = xlwt.Workbook()
    style1 = xlwt.easyxf(num_format_str='D-MMM-YY')
    ws = wb.add_sheet('Bien')
    for oo in av:
            for q,i in zip(ba,oo) :
                if isinstance(i,datetime.datetime) == True:
                    ws.write(v, q, i,style1)
                elif isinstance(i,datetime.date) == True:
                    ws.write(v, q, i,style1) 
                else:
                    ws.write(v, q, i)
            v=v+1
    #wb.save('C:/Users/user/Downloads/Telegram Desktop/BienClient.xls')


def date_(floa,date):
        try:
            v=datetime.datetime(*xlrd.xldate_as_tuple(floa,date))
            return v
        except:
            return floa

def lient(loc):
    anomalie=[]
    anomalie1=[]
    Bien1=[]
    Bien=[]
    a=[]
    b=[]
    p=0
    wb_obj = openpyxl.load_workbook(loc,data_only=True)
    sheet=wb_obj.active
    if sheet["W"][0].value!='CODE POSTAL'and sheet["AB"][0].value!='Tel principal client'and sheet["J"][0].value!='Ref code client- service comptabilité'and sheet["P"][0].value!='Fonction responsable':
                return False
    for i in range(0,sheet.max_row):
            if sheet["J"][i].value!='PROSPECT':
                v=sheet["M"][i].value
                
                if type(v)==int: 
                    if v  not in a:
                        cli=Client.query.filter_by(reference=int(sheet["M"][i].value)).first()
                        a.append(v)
                        if cli is None:
                            create=Client()
                            
                            Histo=Client_History()
                            
                            create.reference=int(sheet["M"][i].value)
                            create.TYPE=regex1(sheet["P"][i].value,'str')
                            create.societe=regex1(sheet["Q"][i].value,'str')
                            create.enseigne=regex1(sheet["R"][i].value,'str')
                            create.titre=regex1(sheet["N"][i].value,'str')
                            create.nom=regex1(sheet["O"][i].value,'str1')
                            numero=regex1(sheet["AB"][i].value,'int')
                            if numero == 'Failed':
                                reason='erreur  de numero dans la colonne  Tel principal client ,veuillez verifier toute colonne avant d"envoyer'
                                anomalie.append([sheet["A"][i].value,sheet["B"][i].value,sheet["C"][i].value,
                                sheet["D"][i].value,sheet["E"][i].value,sheet["F"][i].value,sheet["G"][i].value,
                                sheet["H"][i].value,sheet["I"][i].value,sheet["J"][i].value,sheet["K"][i].value,
                                sheet["L"][i].value,sheet["M"][i].value,sheet["N"][i].value,sheet["O"][i].value,
                                sheet["P"][i].value,sheet["Q"][i].value,sheet["R"][i].value,sheet["S"][i].value,
                                sheet["T"][i].value,sheet["U"][i].value,sheet["V"][i].value,sheet["W"][i].value,
                                sheet["X"][i].value,sheet["Y"][i].value,sheet["Z"][i].value,sheet["AA"][i].value
                                ,sheet["AB"][i].value,sheet["AC"][i].value,reason])
                                continue
                            else:
                                create.numero=numero
                            
                            siret=regex1(sheet["S"][i].value,'int')
                            if siret == 'Failed':
                                reason='erreur  de numero dans la colonne  Siret ,veuillez verifier toute colonne avant d"envoyer'
                                anomalie.append([sheet["A"][i].value,sheet["B"][i].value,sheet["C"][i].value,
                                sheet["D"][i].value,sheet["E"][i].value,sheet["F"][i].value,sheet["G"][i].value,
                                sheet["H"][i].value,sheet["I"][i].value,sheet["J"][i].value,sheet["K"][i].value,
                                sheet["L"][i].value,sheet["M"][i].value,sheet["N"][i].value,sheet["O"][i].value,
                                sheet["P"][i].value,sheet["Q"][i].value,sheet["R"][i].value,sheet["S"][i].value,
                                sheet["T"][i].value,sheet["U"][i].value,sheet["V"][i].value,sheet["W"][i].value,
                                sheet["X"][i].value,sheet["Y"][i].value,sheet["Z"][i].value,sheet["AA"][i].value
                                ,sheet["AB"][i].value,sheet["AC"][i].value,reason])
                                continue
                            else:
                                create.siret=siret
                            create.email=regex1(sheet["AC"][i].value,'str1')
                            date_creation=regex1(sheet["I"][i].value,'date')
                            if date_creation == False:
                                reason='erreur  de numero dans la colonne  date creation ,veuillez verifier toute colonne avant d"envoyer'
                                anomalie.append([sheet["A"][i].value,sheet["B"][i].value,sheet["C"][i].value,
                                sheet["D"][i].value,sheet["E"][i].value,sheet["F"][i].value,sheet["G"][i].value,
                                sheet["H"][i].value,sheet["I"][i].value,sheet["J"][i].value,sheet["K"][i].value,
                                sheet["L"][i].value,sheet["M"][i].value,sheet["N"][i].value,sheet["O"][i].value,
                                sheet["P"][i].value,sheet["Q"][i].value,sheet["R"][i].value,sheet["S"][i].value,
                                sheet["T"][i].value,sheet["U"][i].value,sheet["V"][i].value,sheet["W"][i].value,
                                sheet["X"][i].value,sheet["Y"][i].value,sheet["Z"][i].value,sheet["AA"][i].value
                                ,sheet["AB"][i].value,sheet["AC"][i].value,reason])
                                continue
                            else:
                                create.date_creation=date_creation
                            
                            Histo.adresse1=regex1(sheet["U"][i].value,'str1')
                            Histo.adresse2=regex1(sheet["V"][i].value,'str1')
                            Histo.etat_client=regex1(sheet["J"][i].value,'str')
                            cp=regex1(sheet["W"][i].value,'int')
                            if cp == 'Failed':
                                reason='erreur  de numero dans la colonne  code postal ,veuillez verifier toute colonne avant d"envoyer'
                                anomalie.append([sheet["A"][i].value,sheet["B"][i].value,sheet["C"][i].value,
                                sheet["D"][i].value,sheet["E"][i].value,sheet["F"][i].value,sheet["G"][i].value,
                                sheet["H"][i].value,sheet["I"][i].value,sheet["J"][i].value,sheet["K"][i].value,
                                sheet["L"][i].value,sheet["M"][i].value,sheet["N"][i].value,sheet["O"][i].value,
                                sheet["P"][i].value,sheet["Q"][i].value,sheet["R"][i].value,sheet["S"][i].value,
                                sheet["T"][i].value,sheet["U"][i].value,sheet["V"][i].value,sheet["W"][i].value,
                                sheet["X"][i].value,sheet["Y"][i].value,sheet["Z"][i].value,sheet["AA"][i].value
                                ,sheet["AB"][i].value,sheet["AC"][i].value,reason])
                                continue   
                            else:
                                Histo.cp=cp
                            Histo.ville=regex1(sheet["X"][i].value,'str1')
                            Histo.pays='France'
                            Histo.login_extranet=regex1(sheet["C"][i].value,'str1')
                            Histo.mpd_extranet=regex1(sheet["D"][i].value,'str1')
                            db.session.add(create)
                            db.session.add(Histo)
                            db.session.commit()
                            Histo.client_id=create.id
                            db.session.commit()
                            Bien.append([sheet["A"][i].value,sheet["B"][i].value,sheet["C"][i].value,
                                sheet["D"][i].value,sheet["E"][i].value,sheet["F"][i].value,sheet["G"][i].value,
                                sheet["H"][i].value,sheet["I"][i].value,sheet["J"][i].value,sheet["K"][i].value,
                                sheet["L"][i].value,sheet["M"][i].value,sheet["N"][i].value,sheet["O"][i].value,
                                sheet["P"][i].value,sheet["Q"][i].value,sheet["R"][i].value,sheet["S"][i].value,
                                sheet["T"][i].value,sheet["U"][i].value,sheet["V"][i].value,sheet["W"][i].value,
                                sheet["X"][i].value,sheet["Y"][i].value,sheet["Z"][i].value,sheet["AA"][i].value
                                ,sheet["AB"][i].value,sheet["AC"][i].value])
                        else:
                            print("Anomalie doublon")
                    else:
                        reason="erreur  de doublon dans la colonne  reference,veuillez verifier toute colonne avant d'envoyer"
                        anomalie.append([sheet["A"][i].value,sheet["B"][i].value,sheet["C"][i].value,
                        sheet["D"][i].value,sheet["E"][i].value,sheet["F"][i].value,sheet["G"][i].value,
                        sheet["H"][i].value,sheet["I"][i].value,sheet["J"][i].value,sheet["K"][i].value,
                        sheet["L"][i].value,sheet["M"][i].value,sheet["N"][i].value,sheet["O"][i].value,
                        sheet["P"][i].value,sheet["Q"][i].value,sheet["R"][i].value,sheet["S"][i].value,
                        sheet["T"][i].value,sheet["U"][i].value,sheet["V"][i].value,sheet["W"][i].value,
                        sheet["X"][i].value,sheet["Y"][i].value,sheet["Z"][i].value,sheet["AA"][i].value
                        ,sheet["AB"][i].value,sheet["AC"][i].value,reason])
                else:
                    anomalie.append([sheet["A"][i].value,sheet["B"][i].value,sheet["C"][i].value,
                    sheet["D"][i].value,sheet["E"][i].value,sheet["F"][i].value,sheet["G"][i].value,
                    sheet["H"][i].value,sheet["I"][i].value,sheet["J"][i].value,sheet["K"][i].value,
                    sheet["L"][i].value,sheet["M"][i].value,sheet["N"][i].value,sheet["O"][i].value,
                    sheet["P"][i].value,sheet["Q"][i].value,sheet["R"][i].value,sheet["S"][i].value,
                    sheet["T"][i].value,sheet["U"][i].value,sheet["V"][i].value,sheet["W"][i].value,
                    sheet["X"][i].value,sheet["Y"][i].value,sheet["Z"][i].value,sheet["AA"][i].value
                    ,sheet["AB"][i].value,sheet["AC"][i].value,"erreur dans la colonne  reference,veuillez verifier toute colonne avant d'envoyer"])
            if  sheet["J"][i].value =='PROSPECT':
                    v=sheet["L"][i].value 
                    if type(v)==int: 
                        if v  not in b:
                            cli=prospect.query.filter_by(reference=int(sheet["L"][i].value)).first()
                            b.append(v)
                            if cli is None:
                                create=prospect()
                                
                                Histo=prospect_History()
                                
                                create.reference=int(sheet["L"][i].value)
                                create.TYPE=regex1(sheet["P"][i].value,'str')
                                create.societe=regex1(sheet["Q"][i].value,'str')
                                create.enseigne=regex1(sheet["R"][i].value,'str')
                                create.titre=regex1(sheet["N"][i].value,'str')
                                create.nom=regex1(sheet["O"][i].value,'str1')
                                numero=regex1(sheet["AB"][i].value,'int')
                                if numero == False:
                                    reason='erreur  de numero dans la colonne  Tel principal client,veuillez verifier toute colonne avant d"envoyer'
                                    anomalie1.append([sheet["A"][i].value,sheet["B"][i].value,sheet["C"][i].value,
                                    sheet["D"][i].value,sheet["E"][i].value,sheet["F"][i].value,sheet["G"][i].value,
                                    sheet["H"][i].value,sheet["I"][i].value,sheet["J"][i].value,sheet["K"][i].value,
                                    sheet["L"][i].value,sheet["M"][i].value,sheet["N"][i].value,sheet["O"][i].value,
                                    sheet["P"][i].value,sheet["Q"][i].value,sheet["R"][i].value,sheet["S"][i].value,
                                    sheet["T"][i].value,sheet["U"][i].value,sheet["V"][i].value,sheet["W"][i].value,
                                    sheet["X"][i].value,sheet["Y"][i].value,sheet["Z"][i].value,sheet["AA"][i].value
                                    ,sheet["AB"][i].value,sheet["AC"][i].value,reason])
                                    continue
                                else:
                                    create.numero=numero
                                
                                siret=regex1(sheet["S"][i].value,'int')
                                if siret == False:
                                    reason='erreur  de numero dans la colonne  Siret,veuillez verifier toute colonne avant d"envoyer'
                                    anomalie1.append([sheet["A"][i].value,sheet["B"][i].value,sheet["C"][i].value,
                                    sheet["D"][i].value,sheet["E"][i].value,sheet["F"][i].value,sheet["G"][i].value,
                                    sheet["H"][i].value,sheet["I"][i].value,sheet["J"][i].value,sheet["K"][i].value,
                                    sheet["L"][i].value,sheet["M"][i].value,sheet["N"][i].value,sheet["O"][i].value,
                                    sheet["P"][i].value,sheet["Q"][i].value,sheet["R"][i].value,sheet["S"][i].value,
                                    sheet["T"][i].value,sheet["U"][i].value,sheet["V"][i].value,sheet["W"][i].value,
                                    sheet["X"][i].value,sheet["Y"][i].value,sheet["Z"][i].value,sheet["AA"][i].value
                                    ,sheet["AB"][i].value,sheet["AC"][i].value,reason])
                                    continue
                                else:
                                    create.siret=siret
                                create.email=regex1(sheet["AC"][i].value,'str1')
                                date_creation=regex1(sheet["I"][i].value,'date')
                                if date_creation == False:
                                    reason='erreur  de numero dans la colonne  date entree ,veuillez verifier toute colonne avant d"envoyer'
                                    anomalie1.append([sheet["A"][i].value,sheet["B"][i].value,sheet["C"][i].value,
                                    sheet["D"][i].value,sheet["E"][i].value,sheet["F"][i].value,sheet["G"][i].value,
                                    sheet["H"][i].value,sheet["I"][i].value,sheet["J"][i].value,sheet["K"][i].value,
                                    sheet["L"][i].value,sheet["M"][i].value,sheet["N"][i].value,sheet["O"][i].value,
                                    sheet["P"][i].value,sheet["Q"][i].value,sheet["R"][i].value,sheet["S"][i].value,
                                    sheet["T"][i].value,sheet["U"][i].value,sheet["V"][i].value,sheet["W"][i].value,
                                    sheet["X"][i].value,sheet["Y"][i].value,sheet["Z"][i].value,sheet["AA"][i].value
                                    ,sheet["AB"][i].value,sheet["AC"][i].value,reason])
                                    continue
                                else:
                                    create.date_creation=date_creation
                                
                                Histo.adresse1=regex1(sheet["U"][i].value,'str1')
                                Histo.adresse2=regex1(sheet["V"][i].value,'str1')
                                Histo.etat_client=regex1(sheet["J"][i].value,'str')
                                cp=regex1(sheet["W"][i].value,'int')
                                if cp == False:
                                    reason='erreur  de numero dans la colonne  code postal ,veuillez verifier toute colonne avant d"envoyer'
                                    anomalie1.append([sheet["A"][i].value,sheet["B"][i].value,sheet["C"][i].value,
                                    sheet["D"][i].value,sheet["E"][i].value,sheet["F"][i].value,sheet["G"][i].value,
                                    sheet["H"][i].value,sheet["I"][i].value,sheet["J"][i].value,sheet["K"][i].value,
                                    sheet["L"][i].value,sheet["M"][i].value,sheet["N"][i].value,sheet["O"][i].value,
                                    sheet["P"][i].value,sheet["Q"][i].value,sheet["R"][i].value,sheet["S"][i].value,
                                    sheet["T"][i].value,sheet["U"][i].value,sheet["V"][i].value,sheet["W"][i].value,
                                    sheet["X"][i].value,sheet["Y"][i].value,sheet["Z"][i].value,sheet["AA"][i].value
                                    ,sheet["AB"][i].value,sheet["AC"][i].value,reason])
                                    continue   
                                else:
                                    Histo.cp=cp
                                Histo.ville=regex1(sheet["X"][i].value,'str1')
                                Histo.pays='France'
                                Histo.login_extranet=regex1(sheet["C"][i].value,'str1')
                                Histo.mpd_extranet=regex1(sheet["D"][i].value,'str1')
                                db.session.add(create)
                                db.session.add(Histo)
                                db.session.commit()
                                Histo.prospect=create.id
                                db.session.commit()
                                Bien1.append([sheet["A"][i].value,sheet["B"][i].value,sheet["C"][i].value,
                                sheet["D"][i].value,sheet["E"][i].value,sheet["F"][i].value,sheet["G"][i].value,
                                sheet["H"][i].value,sheet["I"][i].value,sheet["J"][i].value,sheet["K"][i].value,
                                sheet["L"][i].value,sheet["M"][i].value,sheet["N"][i].value,sheet["O"][i].value,
                                sheet["P"][i].value,sheet["Q"][i].value,sheet["R"][i].value,sheet["S"][i].value,
                                sheet["T"][i].value,sheet["U"][i].value,sheet["V"][i].value,sheet["W"][i].value,
                                sheet["X"][i].value,sheet["Y"][i].value,sheet["Z"][i].value,sheet["AA"][i].value
                                ,sheet["AB"][i].value,sheet["AC"][i].value])
                            else:
                                print("Anomalie doublon")
                        else:
                            reason="Anomalie doublon sur la reference"
                            anomalie1.append([sheet["A"][i].value,sheet["B"][i].value,sheet["C"][i].value,
                            sheet["D"][i].value,sheet["E"][i].value,sheet["F"][i].value,sheet["G"][i].value,
                            sheet["H"][i].value,sheet["I"][i].value,sheet["J"][i].value,sheet["K"][i].value,
                            sheet["L"][i].value,sheet["M"][i].value,sheet["N"][i].value,sheet["O"][i].value,
                            sheet["P"][i].value,sheet["Q"][i].value,sheet["R"][i].value,sheet["S"][i].value,
                            sheet["T"][i].value,sheet["U"][i].value,sheet["V"][i].value,sheet["W"][i].value,
                            sheet["X"][i].value,sheet["Y"][i].value,sheet["Z"][i].value,sheet["AA"][i].value
                            ,sheet["AB"][i].value,sheet["AC"][i].value,reason])
                    else:
                        anomalie1.append([sheet["A"][i].value,sheet["B"][i].value,sheet["C"][i].value,
                        sheet["D"][i].value,sheet["E"][i].value,sheet["F"][i].value,sheet["G"][i].value,
                        sheet["H"][i].value,sheet["I"][i].value,sheet["J"][i].value,sheet["K"][i].value,
                        sheet["L"][i].value,sheet["M"][i].value,sheet["N"][i].value,sheet["O"][i].value,
                        sheet["P"][i].value,sheet["Q"][i].value,sheet["R"][i].value,sheet["S"][i].value,
                        sheet["T"][i].value,sheet["U"][i].value,sheet["V"][i].value,sheet["W"][i].value,
                        sheet["X"][i].value,sheet["Y"][i].value,sheet["Z"][i].value,sheet["AA"][i].value
                        ,sheet["AB"][i].value,sheet["AC"][i].value,"erreur dans la colonne  reference,veuillez verifier toute colonne avant d'envoyer"])
    if anomalie == []:
        print(anomalie)
    else:
        failed(anomalie)
    if anomalie1 == []:
        print(anomalie1)
    else:
        failed1(anomalie1)
    if Bien == []:
        print(Bien)
    else:
        good2(Bien)
    if Bien1 == []:
        print(Bien1)
    else:
        good1(Bien1)



