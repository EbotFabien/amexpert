from Database_project.project.data_base_ import db
from Database_project.project.data_base_ import bcrypt
from Database_project.project.data_base_.Models import Tarifs,Mission,Client,Expert,Client_History,prospect,prospect_History,Expert_History,Tarif_base,suivi_client,suivi_prospect
import xlrd,xlwt
import openpyxl
import pandas as pd
from sqlalchemy import or_, and_
import datetime

def checktarif(a,p,c):
    taf_base=Tarifs.query.filter_by(reference_client = int(c)).first()
    if type(a) == int:
            v=float(a)
            if p =="edl_prix_std":
                taf_base.edl_prix_std=v
            if p =="edl_appt_prix_f1":
                taf_base.edl_appt_prix_f1=v
            if p =="edl_appt_prix_f2":
                taf_base.edl_appt_prix_f2=v
            if p =="edl_appt_prix_f3":
                taf_base.edl_appt_prix_f3=v
            if p =="edl_appt_prix_f4":
                taf_base.edl_appt_prix_f4=v
            if p =="edl_appt_prix_f5":
                taf_base.edl_appt_prix_f5=v
            if p =="edl_appt_prix_f6":
                taf_base.edl_appt_prix_f6=v
            if p =="edl_pav_villa_prix_t1":
                taf_base.edl_pav_villa_prix_t1=v
            if p =="edl_pav_villa_prix_t2":
                taf_base.edl_pav_villa_prix_t2=v
            if p =="edl_pav_villa_prix_t3":
                taf_base.edl_pav_villa_prix_t3=v
            if p =="edl_pav_villa_prix_t4":
                taf_base.edl_pav_villa_prix_t4=v
            if p =="edl_pav_villa_prix_t5":
                taf_base.edl_pav_villa_prix_t5=v
            if p =="edl_pav_villa_prix_t6":
                taf_base.edl_pav_villa_prix_t6=v
            if p =="edl_pav_villa_prix_t7":
                taf_base.edl_pav_villa_prix_t7=v
            if p =="edl_pav_villa_prix_t8":
                taf_base.edl_pav_villa_prix_t8=v
            if p =="chif_appt_prix_stu":
                taf_base.chif_appt_prix_stu=v
            if p =="chif_appt_prix_f1":
                taf_base.chif_appt_prix_f1=v
            if p =="chif_appt_prix_f2":
                taf_base.chif_appt_prix_f2=v
            if p =="chif_appt_prix_f3":
                taf_base.chif_appt_prix_f3=v
            if p =="chif_appt_prix_f4":
                taf_base.chif_appt_prix_f4=v
            if p =="chif_appt_prix_f5":
                taf_base.chif_appt_prix_f5=v
            if p =="chif_pav_villa_prix_t1":
                taf_base.chif_pav_villa_prix_t1=v
            if p =="chif_pav_villa_prix_t2":
                taf_base.chif_pav_villa_prix_t2=v
            if p =="chif_pav_villa_prix_t3":
                taf_base.chif_pav_villa_prix_t3=v
            if p =="chif_pav_villa_prix_t4":
                taf_base.chif_pav_villa_prix_t4=v
            if p =="chif_pav_villa_prix_t5":
                taf_base.chif_pav_villa_prix_t5=v
            if p =="chif_pav_villa_prix_t6":
                taf_base.chif_pav_villa_prix_t6=v
            if p =="chif_pav_villa_prix_t7":
                taf_base.chif_pav_villa_prix_t7=v
            if p =="chif_pav_villa_prix_t8":
                taf_base.chif_pav_villa_prix_t8=v
                
            db.session.commit()
    if type(a) == str:
              
        try:
            v=a[0:a.index('€')]
            try:
                a=float(v)
                if p =="edl_prix_std":
                    taf_base.edl_prix_std=v
                if p =="edl_appt_prix_f1":
                    taf_base.edl_appt_prix_f1=v
                if p =="edl_appt_prix_f2":
                    taf_base.edl_appt_prix_f2=v
                if p =="edl_appt_prix_f3":
                    taf_base.edl_appt_prix_f3=v
                if p =="edl_appt_prix_f4":
                    taf_base.edl_appt_prix_f4=v
                if p =="edl_appt_prix_f5":
                    taf_base.edl_appt_prix_f5=v
                if p =="edl_appt_prix_f6":
                    taf_base.edl_appt_prix_f6=v
                if p =="edl_pav_villa_prix_t1":
                    taf_base.edl_pav_villa_prix_t1=v
                if p =="edl_pav_villa_prix_t2":
                    taf_base.edl_pav_villa_prix_t2=v
                if p =="edl_pav_villa_prix_t3":
                    taf_base.edl_pav_villa_prix_t3=v
                if p =="edl_pav_villa_prix_t4":
                    taf_base.edl_pav_villa_prix_t4=v
                if p =="edl_pav_villa_prix_t5":
                    taf_base.edl_pav_villa_prix_t5=v
                if p =="edl_pav_villa_prix_t6":
                    taf_base.edl_pav_villa_prix_t6=v
                if p =="edl_pav_villa_prix_t7":
                    taf_base.edl_pav_villa_prix_t7=v
                if p =="edl_pav_villa_prix_t8":
                    taf_base.edl_pav_villa_prix_t8=v
                if p =="chif_appt_prix_stu":
                    taf_base.chif_appt_prix_stu=v
                if p =="chif_appt_prix_f1":
                    taf_base.chif_appt_prix_f1=v
                if p =="chif_appt_prix_f2":
                    taf_base.chif_appt_prix_f2=v
                if p =="chif_appt_prix_f3":
                    taf_base.chif_appt_prix_f3=v
                if p =="chif_appt_prix_f4":
                    taf_base.chif_appt_prix_f4=v
                if p =="chif_appt_prix_f5":
                    taf_base.chif_appt_prix_f5=v
                if p =="chif_pav_villa_prix_t1":
                    taf_base.chif_pav_villa_prix_t1=v
                if p =="chif_pav_villa_prix_t2":
                    taf_base.chif_pav_villa_prix_t2=v
                if p =="chif_pav_villa_prix_t3":
                    taf_base.chif_pav_villa_prix_t3=v
                if p =="chif_pav_villa_prix_t4":
                    taf_base.chif_pav_villa_prix_t4=v
                if p =="chif_pav_villa_prix_t5":
                    taf_base.chif_pav_villa_prix_t5=v
                if p =="chif_pav_villa_prix_t6":
                    taf_base.chif_pav_villa_prix_t6=v
                if p =="chif_pav_villa_prix_t7":
                    taf_base.chif_pav_villa_prix_t7=v
                if p =="chif_pav_villa_prix_t8":
                    taf_base.chif_pav_villa_prix_t8=v
                db.session.commit() 
            except:
                f=v[2:]
                a=[]
                for word in f.split():
                    if word.isdigit():
                        a.append(int(word))
                vam=str(a)
                v=vam[1:-1]
                if p =="edl_prix_std":
                    taf_base.edl_prix_std=v
                if p =="edl_appt_prix_f1":
                    taf_base.edl_appt_prix_f1=v
                if p =="edl_appt_prix_f2":
                    taf_base.edl_appt_prix_f2=v
                if p =="edl_appt_prix_f3":
                    taf_base.edl_appt_prix_f3=v
                if p =="edl_appt_prix_f4":
                    taf_base.edl_appt_prix_f4=v
                if p =="edl_appt_prix_f5":
                    taf_base.edl_appt_prix_f5=v
                if p =="edl_appt_prix_f6":
                    taf_base.edl_appt_prix_f6=v
                if p =="edl_pav_villa_prix_t1":
                    taf_base.edl_pav_villa_prix_t1=v
                if p =="edl_pav_villa_prix_t2":
                    taf_base.edl_pav_villa_prix_t2=v
                if p =="edl_pav_villa_prix_t3":
                    taf_base.edl_pav_villa_prix_t3=v
                if p =="edl_pav_villa_prix_t4":
                    taf_base.edl_pav_villa_prix_t4=v
                if p =="edl_pav_villa_prix_t5":
                    taf_base.edl_pav_villa_prix_t5=v
                if p =="edl_pav_villa_prix_t6":
                    taf_base.edl_pav_villa_prix_t6=v
                if p =="edl_pav_villa_prix_t7":
                    taf_base.edl_pav_villa_prix_t7=v
                if p =="edl_pav_villa_prix_t8":
                    taf_base.edl_pav_villa_prix_t8=v
                if p =="chif_appt_prix_stu":
                    taf_base.chif_appt_prix_stu=v
                if p =="chif_appt_prix_f1":
                    taf_base.chif_appt_prix_f1=v
                if p =="chif_appt_prix_f2":
                    taf_base.chif_appt_prix_f2=v
                if p =="chif_appt_prix_f3":
                    taf_base.chif_appt_prix_f3=v
                if p =="chif_appt_prix_f4":
                    taf_base.chif_appt_prix_f4=v
                if p =="chif_appt_prix_f5":
                    taf_base.chif_appt_prix_f5=v
                if p =="chif_pav_villa_prix_t1":
                    taf_base.chif_pav_villa_prix_t1=v
                if p =="chif_pav_villa_prix_t2":
                    taf_base.chif_pav_villa_prix_t2=v
                if p =="chif_pav_villa_prix_t3":
                    taf_base.chif_pav_villa_prix_t3=v
                if p =="chif_pav_villa_prix_t4":
                    taf_base.chif_pav_villa_prix_t4=v
                if p =="chif_pav_villa_prix_t5":
                    taf_base.chif_pav_villa_prix_t5=v
                if p =="chif_pav_villa_prix_t6":
                    taf_base.chif_pav_villa_prix_t6=v
                if p =="chif_pav_villa_prix_t7":
                    taf_base.chif_pav_villa_prix_t7=v
                if p =="chif_pav_villa_prix_t8":
                    taf_base.chif_pav_villa_prix_t8=v
                db.session.commit()

                

        except:
              try:
                f=int(a)
                v=float(f)
                if p =="edl_prix_std":
                    taf_base.edl_prix_std=v
                if p =="edl_appt_prix_f1":
                    taf_base.edl_appt_prix_f1=v
                if p =="edl_appt_prix_f2":
                    taf_base.edl_appt_prix_f2=v
                if p =="edl_appt_prix_f3":
                    taf_base.edl_appt_prix_f3=v
                if p =="edl_appt_prix_f4":
                    taf_base.edl_appt_prix_f4=v
                if p =="edl_appt_prix_f5":
                    taf_base.edl_appt_prix_f5=v
                if p =="edl_appt_prix_f6":
                    taf_base.edl_appt_prix_f6=v
                if p =="edl_pav_villa_prix_t1":
                    taf_base.edl_pav_villa_prix_t1=v
                if p =="edl_pav_villa_prix_t2":
                    taf_base.edl_pav_villa_prix_t2=v
                if p =="edl_pav_villa_prix_t3":
                    taf_base.edl_pav_villa_prix_t3=v
                if p =="edl_pav_villa_prix_t4":
                    taf_base.edl_pav_villa_prix_t4=v
                if p =="edl_pav_villa_prix_t5":
                    taf_base.edl_pav_villa_prix_t5=v
                if p =="edl_pav_villa_prix_t6":
                    taf_base.edl_pav_villa_prix_t6=v
                if p =="edl_pav_villa_prix_t7":
                    taf_base.edl_pav_villa_prix_t7=v
                if p =="edl_pav_villa_prix_t8":
                    taf_base.edl_pav_villa_prix_t8=v
                if p =="chif_appt_prix_stu":
                    taf_base.chif_appt_prix_stu=v
                if p =="chif_appt_prix_f1":
                    taf_base.chif_appt_prix_f1=v
                if p =="chif_appt_prix_f2":
                    taf_base.chif_appt_prix_f2=v
                if p =="chif_appt_prix_f3":
                    taf_base.chif_appt_prix_f3=v
                if p =="chif_appt_prix_f4":
                    taf_base.chif_appt_prix_f4=v
                if p =="chif_appt_prix_f5":
                    taf_base.chif_appt_prix_f5=v
                if p =="chif_pav_villa_prix_t1":
                    taf_base.chif_pav_villa_prix_t1=v
                if p =="chif_pav_villa_prix_t2":
                    taf_base.chif_pav_villa_prix_t2=v
                if p =="chif_pav_villa_prix_t3":
                    taf_base.chif_pav_villa_prix_t3=v
                if p =="chif_pav_villa_prix_t4":
                    taf_base.chif_pav_villa_prix_t4=v
                if p =="chif_pav_villa_prix_t5":
                    taf_base.chif_pav_villa_prix_t5=v
                if p =="chif_pav_villa_prix_t6":
                    taf_base.chif_pav_villa_prix_t6=v
                if p =="chif_pav_villa_prix_t7":
                    taf_base.chif_pav_villa_prix_t7=v
                if p =="chif_pav_villa_prix_t8":
                    taf_base.chif_pav_villa_prix_t8=v
                db.session.commit()
              except:
                    v=0
                    if p =="edl_prix_std":
                        taf_base.edl_prix_std=v
                    if p =="edl_appt_prix_f1":
                        taf_base.edl_appt_prix_f1=v
                    if p =="edl_appt_prix_f2":
                        taf_base.edl_appt_prix_f2=v
                    if p =="edl_appt_prix_f3":
                        taf_base.edl_appt_prix_f3=v
                    if p =="edl_appt_prix_f4":
                        taf_base.edl_appt_prix_f4=v
                    if p =="edl_appt_prix_f5":
                        taf_base.edl_appt_prix_f5=v
                    if p =="edl_appt_prix_f6":
                        taf_base.edl_appt_prix_f6=v
                    if p =="edl_pav_villa_prix_t1":
                        taf_base.edl_pav_villa_prix_t1=v
                    if p =="edl_pav_villa_prix_t2":
                        taf_base.edl_pav_villa_prix_t2=v
                    if p =="edl_pav_villa_prix_t3":
                        taf_base.edl_pav_villa_prix_t3=v
                    if p =="edl_pav_villa_prix_t4":
                        taf_base.edl_pav_villa_prix_t4=v
                    if p =="edl_pav_villa_prix_t5":
                        taf_base.edl_pav_villa_prix_t5=v
                    if p =="edl_pav_villa_prix_t6":
                        taf_base.edl_pav_villa_prix_t6=v
                    if p =="edl_pav_villa_prix_t7":
                        taf_base.edl_pav_villa_prix_t7=v
                    if p =="edl_pav_villa_prix_t8":
                        taf_base.edl_pav_villa_prix_t8=v
                    if p =="chif_appt_prix_stu":
                        taf_base.chif_appt_prix_stu=v
                    if p =="chif_appt_prix_f1":
                        taf_base.chif_appt_prix_f1=v
                    if p =="chif_appt_prix_f2":
                        taf_base.chif_appt_prix_f2=v
                    if p =="chif_appt_prix_f3":
                        taf_base.chif_appt_prix_f3=v
                    if p =="chif_appt_prix_f4":
                        taf_base.chif_appt_prix_f4=v
                    if p =="chif_appt_prix_f5":
                        taf_base.chif_appt_prix_f5=v
                    if p =="chif_pav_villa_prix_t1":
                        taf_base.chif_pav_villa_prix_t1=v
                    if p =="chif_pav_villa_prix_t2":
                        taf_base.chif_pav_villa_prix_t2=v
                    if p =="chif_pav_villa_prix_t3":
                        taf_base.chif_pav_villa_prix_t3=v
                    if p =="chif_pav_villa_prix_t4":
                        taf_base.chif_pav_villa_prix_t4=v
                    if p =="chif_pav_villa_prix_t5":
                        taf_base.chif_pav_villa_prix_t5=v
                    if p =="chif_pav_villa_prix_t6":
                        taf_base.chif_pav_villa_prix_t6=v
                    if p =="chif_pav_villa_prix_t7":
                        taf_base.chif_pav_villa_prix_t7=v
                    if p =="chif_pav_villa_prix_t8":
                        taf_base.chif_pav_villa_prix_t8=v

                    db.session.commit()
    if type(a) == float:
        try:
                f=int(a)
                v=float(f)
                if p =="edl_prix_std":
                    taf_base.edl_prix_std=v
                if p =="edl_appt_prix_f1":
                    taf_base.edl_appt_prix_f1=v
                if p =="edl_appt_prix_f2":
                    taf_base.edl_appt_prix_f2=v
                if p =="edl_appt_prix_f3":
                    taf_base.edl_appt_prix_f3=v
                if p =="edl_appt_prix_f4":
                    taf_base.edl_appt_prix_f4=v
                if p =="edl_appt_prix_f5":
                    taf_base.edl_appt_prix_f5=v
                if p =="edl_appt_prix_f6":
                    taf_base.edl_appt_prix_f6=v
                if p =="edl_pav_villa_prix_t1":
                    taf_base.edl_pav_villa_prix_t1=v
                if p =="edl_pav_villa_prix_t2":
                    taf_base.edl_pav_villa_prix_t2=v
                if p =="edl_pav_villa_prix_t3":
                    taf_base.edl_pav_villa_prix_t3=v
                if p =="edl_pav_villa_prix_t4":
                    taf_base.edl_pav_villa_prix_t4=v
                if p =="edl_pav_villa_prix_t5":
                    taf_base.edl_pav_villa_prix_t5=v
                if p =="edl_pav_villa_prix_t6":
                    taf_base.edl_pav_villa_prix_t6=v
                if p =="edl_pav_villa_prix_t7":
                    taf_base.edl_pav_villa_prix_t7=v
                if p =="edl_pav_villa_prix_t8":
                    taf_base.edl_pav_villa_prix_t8=v
                if p =="chif_appt_prix_stu":
                    taf_base.chif_appt_prix_stu=v
                if p =="chif_appt_prix_f1":
                    taf_base.chif_appt_prix_f1=v
                if p =="chif_appt_prix_f2":
                    taf_base.chif_appt_prix_f2=v
                if p =="chif_appt_prix_f3":
                    taf_base.chif_appt_prix_f3=v
                if p =="chif_appt_prix_f4":
                    taf_base.chif_appt_prix_f4=v
                if p =="chif_appt_prix_f5":
                    taf_base.chif_appt_prix_f5=v
                if p =="chif_pav_villa_prix_t1":
                    taf_base.chif_pav_villa_prix_t1=v
                if p =="chif_pav_villa_prix_t2":
                    taf_base.chif_pav_villa_prix_t2=v
                if p =="chif_pav_villa_prix_t3":
                    taf_base.chif_pav_villa_prix_t3=v
                if p =="chif_pav_villa_prix_t4":
                    taf_base.chif_pav_villa_prix_t4=v
                if p =="chif_pav_villa_prix_t5":
                    taf_base.chif_pav_villa_prix_t5=v
                if p =="chif_pav_villa_prix_t6":
                    taf_base.chif_pav_villa_prix_t6=v
                if p =="chif_pav_villa_prix_t7":
                    taf_base.chif_pav_villa_prix_t7=v
                if p =="chif_pav_villa_prix_t8":
                    taf_base.chif_pav_villa_prix_t8=v
                db.session.commit()
        except:
            v=float(0)
            if p =="edl_prix_std":
                taf_base.edl_prix_std=v
            if p =="edl_appt_prix_f1":
                taf_base.edl_appt_prix_f1=v
            if p =="edl_appt_prix_f2":
                taf_base.edl_appt_prix_f2=v
            if p =="edl_appt_prix_f3":
                taf_base.edl_appt_prix_f3=v
            if p =="edl_appt_prix_f4":
                taf_base.edl_appt_prix_f4=v
            if p =="edl_appt_prix_f5":
                taf_base.edl_appt_prix_f5=v
            if p =="edl_appt_prix_f6":
                taf_base.edl_appt_prix_f6=v
            if p =="edl_pav_villa_prix_t1":
                taf_base.edl_pav_villa_prix_t1=v
            if p =="edl_pav_villa_prix_t2":
                taf_base.edl_pav_villa_prix_t2=v
            if p =="edl_pav_villa_prix_t3":
                taf_base.edl_pav_villa_prix_t3=v
            if p =="edl_pav_villa_prix_t4":
                taf_base.edl_pav_villa_prix_t4=v
            if p =="edl_pav_villa_prix_t5":
                taf_base.edl_pav_villa_prix_t5=v
            if p =="edl_pav_villa_prix_t6":
                taf_base.edl_pav_villa_prix_t6=v
            if p =="edl_pav_villa_prix_t7":
                taf_base.edl_pav_villa_prix_t7=v
            if p =="edl_pav_villa_prix_t8":
                taf_base.edl_pav_villa_prix_t8=v
            if p =="chif_appt_prix_stu":
                taf_base.chif_appt_prix_stu=v
            if p =="chif_appt_prix_f1":
                taf_base.chif_appt_prix_f1=v
            if p =="chif_appt_prix_f2":
                taf_base.chif_appt_prix_f2=v
            if p =="chif_appt_prix_f3":
                taf_base.chif_appt_prix_f3=v
            if p =="chif_appt_prix_f4":
                taf_base.chif_appt_prix_f4=v
            if p =="chif_appt_prix_f5":
                taf_base.chif_appt_prix_f5=v
            if p =="chif_pav_villa_prix_t1":
                taf_base.chif_pav_villa_prix_t1=v
            if p =="chif_pav_villa_prix_t2":
                taf_base.chif_pav_villa_prix_t2=v
            if p =="chif_pav_villa_prix_t3":
                taf_base.chif_pav_villa_prix_t3=v
            if p =="chif_pav_villa_prix_t4":
                taf_base.chif_pav_villa_prix_t4=v
            if p =="chif_pav_villa_prix_t5":
                taf_base.chif_pav_villa_prix_t5=v
            if p =="chif_pav_villa_prix_t6":
                taf_base.chif_pav_villa_prix_t6=v
            if p =="chif_pav_villa_prix_t7":
                taf_base.chif_pav_villa_prix_t7=v
            if p =="chif_pav_villa_prix_t8":
                taf_base.chif_pav_villa_prix_t8=v
            db.session.commit()
            
        

def insert_client(loc):
    df = pd.read_excel(loc)
   


    for (ca,a,b,c,d,e,f,g,h,i,j,k,l,m,v,en,n,o,p,q,r,s,t,u,vo,x,y,z,aa,bb,cc,dd,ee,ff,ab,ac,ad,ae,af,ag,ah,ai,aj,ak,al,am,an,ao,ap,aq,ar,As,at,au,av,aw,ax,ay,az,ba,bc,bd,be,bf,bet,bat) in zip(df['Ref code client- service comptabilité'],df['Etat du client']
                                                                            ,df['Référence de la société'],df['Titre'],
                                                                            df['Nom du ou des responsables'],df['Date création du client']
                                                                            ,df['Adresse 1'],df['Complément adresse'],df['CODE POSTAL'],
                                                                            df['Ville'],df['Extranet ID Login'],df['Extranet MDP'],df['Numéro de siret']
                                                                            ,df['AS du Client'],df['Commentaires  suivi client et action à prévoir'],df["Nom de l'enseigne"],df['% Com AS du client'],df['Nom Respon Cell Dev'],
                                                                            df['% Respon Cell Dev'],df['Nom agent Cell Dev'],df['% Agent Cell Dev']
                                                                            ,df['Nom Agent TECH EDELE'],df['% Agent TECH EDELE'],df['Nom Respon Cell Tech']
                                                                            ,df['% Respon Cell Tech'],df['Nom Suiveur Cell Tech'],df['% Suiveur Cell Tech'],
                                                                            df['Nom Respon Cell Planif'],df['% Respon Cell Planif'],df['Nom Suiveur Cell Planif'],
                                                                            df['% Suiveur Cell Planif'],df['Nom Agent saisie Cell Planif'],df['% Agent saisie CEll planif'],
                                                                            df['TAUX DE MAJORATION MEUBLE'],df['EDL_PRIX STD'],df['EDL_APPT / PRIX T1'],df['EDL_APPT / PRIX T2'],
                                                                            df['EDL_APPT / PRIX T3'],df['EDL_APPT / PRIX T4'],df['EDL_APPT / PRIX T5'],df['EDL_APPT / PRIX T6'],
                                                                            df['EDL_PAV VILLA / PRIX T1'],df['EDL_PAV VILLA / PRIX T2'],df['EDL_PAV VILLA / PRIX T3'],
                                                                            df['EDL_PAV VILLA / PRIX T4'],df['EDL_PAV VILLA / PRIX T5'],df['EDL_PAV VILLA / PRIX T6'],
                                                                            df['EDL_PAV VILLA / PRIX T7'],df['EDL_PAV VILLA / PRIX T8'],df['CHIF_APPT / PRIX STU'],
                                                                              df['CHIF_APPT / PRIX T1'],df['CHIF_APPT / PRIX T2'],df['CHIF_APPT / PRIX T3'],
                                                                            df['CHIF_APPT / PRIX T4'],df['CHIF_APPT / PRIX T5'],df['CHIF_PAV VILLA / PRIX T1'],
                                                                            df['CHIF_PAV VILLA / PRIX T2'],df['CHIF_PAV VILLA / PRIX T3'],df['CHIF_PAV VILLA / PRIX T4'],
                                                                            df['CHIF_PAV VILLA / PRIX T5'],df['CHIF_PAV VILLA / PRIX T6'],df['CHIF_PAV VILLA / PRIX T7'],
                                                                            df['CHIF_PAV VILLA / PRIX T8'],df['PRIX AUTRE'],df['Tel principal client'],df['Email de contact général du client']):
                    
                    if str(a) != 'PROSPECT':
                        '''try :
                            cli=Client.query.filter(and_(Client.reference==str(ca),Client.societe==str(b),Client.nom==str(d.lower()))).first()  #check by reference number,societe, and nom
                        except:
                            cli=Client.query.filter(and_(Client.reference==str(ca),Client.societe==str(b))).first() #check by reference number,societe, and nom'''
                        if type(m) != str :
                            r_C=0
                        else:
                            r_C=Expert.query.filter_by(nom=str(m.lower())).first()
                            if r_C is not None:#check keys
                                rC= r_C.id
                            if r_C is None  :
                                rC= 0
                        #if cli is None:
                        try:
                            client=Client(TYPE=str(a),societe=str(b),titre=str(c),nom=str(d.lower()),siret=str(l),numero=str(bet),email=str(bat))#,date_creation=e
                        except:
                            client=Client(TYPE=str(a),societe=str(b),titre=str(c),siret=str(l),numero=str(bet),email=str(bat))#,date_creation=e
                        db.session.add(client)
                        db.session.commit() 
                        if type(ca) != float :
                            client.reference=ca
                            db.session.commit()
                        else:
                            client.reference=''
                            db.session.commit()
                        try:  
                            if type(en) == str:
                                a=str(en.lower()).split()
                                vii=''.join(a)
                                vii=vii.split("-")
                                vii=''.join(vii)
                                client.enseigne =str(vii)
                                db.session.commit()
                        except:
                            print('ei choke')
                        history=Client_History(client_id=client.id,adresse1=str(f),adresse2=str(g),cp=str(h),ville=str(i),
                        login_extranet=str(j),mpd_extranet=str(k),pays="FRANCE",etat_client=str(a))
                        
                        
                        if type(v) == str :#check
                            
                            suivi =suivi_client(client=client.id,responsable=str(rC),commentaire=str(v))
                            db.session.add(suivi)
                        db.session.add(history)
                        db.session.commit()

                        try:
                            r_C=Expert.query.filter_by(nom=str(m.lower())).first()
                            if r_C is not None:#check keys
                                rC= r_C.id
                            if r_C is None  :
                                rC= 0
                        except:
                            rC= 0
                        try:
                            r_r=Expert.query.filter_by(nom=str(o.lower())).first()
                            if r_r is not None:
                                rr= r_r.id
                            if r_r is None  :
                                rr= 0
                        except:
                            rr=0
                        try:
                            dr_a=Expert.query.filter_by(nom=str(q.lower())).first()
                            if dr_a is not None:
                                dra= dr_a.id
                            if dr_a is None  :
                                dra= 0
                        except:
                            dra= 0
                        try:
                            tr_a=Expert.query.filter_by(nom=str(s.lower())).first()
                            if tr_a is not None:
                                tra= tr_a.id
                            if tr_a is None  :
                                tra= 0
                        except:
                            tra=0
                        try:
                            tr_r=Expert.query.filter_by(nom=str(u.lower())).first()
                            if tr_r is not None:
                                trr= tr_r.id
                            if tr_r is None  :
                                trr= 0
                        except:
                            trr= 0
                        try:
                            tr_s=Expert.query.filter_by(nom=str(x.lower())).first()
                            if tr_s is not None:
                                trs= tr_s.id
                            if tr_s is None  :
                                trs= 0
                        except:
                            trs= 0
                        try:
                            pr_s=Expert.query.filter_by(nom=str(z.lower())).first()
                            if pr_s is not None:
                                prs= pr_s.id
                            if pr_s is None  :
                                prs= 0
                        except:
                            prs=0
                        try:
                            pr_si=Expert.query.filter_by(nom=str(bb.lower())).first()
                            if pr_si is not None:
                                prsi= pr_si.id
                            if pr_si is None  :
                                prsi= 0
                        except:
                            prsi= 0
                        try:
                            ra_s=Expert.query.filter_by(nom=str(dd.lower())).first()
                            if ra_s is not None:
                                ras= ra_s.id
                            if ra_s is None  :
                                ras= 0  
                        except:
                            ras= 0    
                        
                        taf_base =Tarifs(reference_client=client.id,
                        referent_as_client=rC,com_as_sur_ca_client=str(n),cell_dev_ref_responsable=rr,
                        com_cell_dev_ref_responsable=str(p),cell_dev_ref_agent=dra,com_cell_dev_ref_agent=str(r),
                        cell_tech_ref_agent=tra,com_cell_tech_Ref_agent=str(t),cell_tech_ref_responsable=trr,
                        com_cell_tech_ref_responsable=str(vo),cell_tech_ref_suiveur=trs,
                        com_cell_tech_ref_suiveur=str(y),cell_planif_ref_responsable=prs,
                        com_cell_planif_ref_responsable=str(aa),cell_planif_ref_suiveur=prsi,
                        com_cell_planif_ref_suiveur=str(cc),cell_planif_ref_agent_saisie=ras,
                        com_cell_planif_ref_agent_saisie=str(ee),taux_meuble=str(ff),prix_autre=str(bf))
                        
                        db.session.add(taf_base)
                        db.session.commit()
                        checktarif(ab,"edl_prix_std",client.id)
                        checktarif(ac,"edl_appt_prix_f1",client.id)
                        checktarif(ad,"edl_appt_prix_f2",client.id)
                        checktarif(ae,"edl_appt_prix_f3",client.id)
                        checktarif(af,"edl_appt_prix_f4",client.id)
                        checktarif(ag,"edl_appt_prix_f5",client.id)
                        checktarif(ah,"edl_appt_prix_f6",client.id)
                        checktarif(ai,"edl_pav_villa_prix_t1",client.id)
                        checktarif(aj,"edl_pav_villa_prix_t2",client.id)
                        checktarif(ak,"edl_pav_villa_prix_t3",client.id)
                        checktarif(al,"edl_pav_villa_prix_t4",client.id)
                        checktarif(am,"edl_pav_villa_prix_t5",client.id)
                        checktarif(an,"edl_pav_villa_prix_t6",client.id)
                        checktarif(ao,"edl_pav_villa_prix_t7",client.id)
                        checktarif(ap,"edl_pav_villa_prix_t8",client.id)
                        checktarif(aq,"chif_appt_prix_stu",client.id)
                        checktarif(ar,"chif_appt_prix_f1",client.id)
                        checktarif(As,"chif_appt_prix_f2",client.id)
                        checktarif(at,"chif_appt_prix_f3",client.id)
                        checktarif(au,"chif_appt_prix_f4",client.id)
                        checktarif(av,"chif_appt_prix_f5",client.id)
                        checktarif(aw,"chif_pav_villa_prix_t1",client.id)
                        checktarif(ax,"chif_pav_villa_prix_t2",client.id)
                        checktarif(ay,"chif_pav_villa_prix_t3",client.id)
                        checktarif(az,"chif_pav_villa_prix_t4",client.id)
                        checktarif(ba,"chif_pav_villa_prix_t5",client.id)
                        checktarif(bc,"chif_pav_villa_prix_t6",client.id)
                        checktarif(bd,"chif_pav_villa_prix_t7",client.id)
                        checktarif(be,"chif_pav_villa_prix_t8",client.id)
                        
                            
                    if str(a) == 'PROSPECT' :
                        
                        '''try:
                            cli=prospect.query.filter(and_(prospect.reference==str(ca),prospect.societe==str(b),prospect.nom==str(d.lower()))).first() #check by reference number,societe, and nom
                        except:
                            cli=prospect.query.filter(and_(prospect.reference==str(ca),prospect.societe==str(b))).first() #check by reference number,societe, and nom
                        if cli is None:'''
                        try:
                            client=prospect(TYPE=a,societe=str(b),titre=str(c),nom=str(d.lower()),siret=str(l),numero=str(bet),email=str(bat))#,date_creation=e
                        except:
                            client=prospect(TYPE=a,societe=str(b),titre=str(c),siret=str(l),numero=str(bet),email=str(bat))#,date_creation=e
                        db.session.add(client)
                        db.session.commit()   
                        if type(ca) != float :
                            client.reference=ca
                            db.session.commit()
                        try:  
                            a=str(en.lower()).split()
                            v=''.join(a)
                            client.enseigne =str(v)
                            db.session.commit()
                        except:
                            print('ei choke')
                        history=prospect_History(prospect=str(client.id),adresse=str(f),cp=str(h),ville=str(i),login_extranet=str(j),mpd_extranet=str(k),pays="France",etat_client=str(a))
                        db.session.add(history)
                        db.session.commit()

                        if type(v) == str :#check
                            
                            suivi =suivi_prospect(client.id,str(rC),str(v))
                            db.session.add(suivi)
                        
                        db.session.commit()

                        #else:
                         #   return 'This data already exist'
                    if  type(ca) != float :
                      cli=Client.query.filter_by(reference=str(ca)).first()
                      if cli is None:
                        if type(m) != str :
                            r_C=0
                        else:
                            r_C=Expert.query.filter_by(nom=str(m.lower())).first()
                            if r_C is not None:#check keys
                                rC= r_C.id
                            if r_C is None  :
                                rC= 0
                        #if cli is None:
                        try:
                            client=Client(TYPE=str(a),societe=str(b),titre=str(c),nom=str(d.lower()),siret=str(l),numero=str(bet),email=str(bat),reference=ca)#,date_creation=e
                        except:
                            client=Client(TYPE=str(a),societe=str(b),titre=str(c),siret=str(l),numero=str(bet),email=str(bat),reference=ca)#,date_creation=e
                        db.session.add(client)
                        db.session.commit() 
                        try:  
                            if type(en) == str:
                                a=str(en.lower()).split()
                                vii=''.join(a)
                                vii=vii.split("-")
                                vii=''.join(vii)
                                client.enseigne =str(vii)
                                db.session.commit()
                        except:
                            print('ei choke')
                        history=Client_History(client_id=client.id,adresse1=str(f),adresse2=str(g),cp=str(h),ville=str(i),
                        login_extranet=str(j),mpd_extranet=str(k),pays="FRANCE",etat_client=str(a))
                        
                        
                        if type(v) == str :#check
                            
                            suivi =suivi_client(client=client.id,responsable=str(rC),commentaire=str(v))
                            db.session.add(suivi)
                        db.session.add(history)
                        db.session.commit()

                        try:
                            r_C=Expert.query.filter_by(nom=str(m.lower())).first()
                            if r_C is not None:#check keys
                                rC= r_C.id
                            if r_C is None  :
                                rC= 0
                        except:
                            rC= 0
                        try:
                            r_r=Expert.query.filter_by(nom=str(o.lower())).first()
                            if r_r is not None:
                                rr= r_r.id
                            if r_r is None  :
                                rr= 0
                        except:
                            rr=0
                        try:
                            dr_a=Expert.query.filter_by(nom=str(q.lower())).first()
                            if dr_a is not None:
                                dra= dr_a.id
                            if dr_a is None  :
                                dra= 0
                        except:
                            dra= 0
                        try:
                            tr_a=Expert.query.filter_by(nom=str(s.lower())).first()
                            if tr_a is not None:
                                tra= tr_a.id
                            if tr_a is None  :
                                tra= 0
                        except:
                            tra=0
                        try:
                            tr_r=Expert.query.filter_by(nom=str(u.lower())).first()
                            if tr_r is not None:
                                trr= tr_r.id
                            if tr_r is None  :
                                trr= 0
                        except:
                            trr= 0
                        try:
                            tr_s=Expert.query.filter_by(nom=str(x.lower())).first()
                            if tr_s is not None:
                                trs= tr_s.id
                            if tr_s is None  :
                                trs= 0
                        except:
                            trs= 0
                        try:
                            pr_s=Expert.query.filter_by(nom=str(z.lower())).first()
                            if pr_s is not None:
                                prs= pr_s.id
                            if pr_s is None  :
                                prs= 0
                        except:
                            prs=0
                        try:
                            pr_si=Expert.query.filter_by(nom=str(bb.lower())).first()
                            if pr_si is not None:
                                prsi= pr_si.id
                            if pr_si is None  :
                                prsi= 0
                        except:
                            prsi= 0
                        try:
                            ra_s=Expert.query.filter_by(nom=str(dd.lower())).first()
                            if ra_s is not None:
                                ras= ra_s.id
                            if ra_s is None  :
                                ras= 0  
                        except:
                            ras= 0    
                        
                        taf_base =Tarifs(reference_client=client.id,
                        referent_as_client=rC,com_as_sur_ca_client=str(n),cell_dev_ref_responsable=rr,
                        com_cell_dev_ref_responsable=str(p),cell_dev_ref_agent=dra,com_cell_dev_ref_agent=str(r),
                        cell_tech_ref_agent=tra,com_cell_tech_Ref_agent=str(t),cell_tech_ref_responsable=trr,
                        com_cell_tech_ref_responsable=str(vo),cell_tech_ref_suiveur=trs,
                        com_cell_tech_ref_suiveur=str(y),cell_planif_ref_responsable=prs,
                        com_cell_planif_ref_responsable=str(aa),cell_planif_ref_suiveur=prsi,
                        com_cell_planif_ref_suiveur=str(cc),cell_planif_ref_agent_saisie=ras,
                        com_cell_planif_ref_agent_saisie=str(ee),taux_meuble=str(ff),prix_autre=str(bf))
                        
                        db.session.add(taf_base)
                        db.session.commit()
                        checktarif(ab,"edl_prix_std",client.id)
                        checktarif(ac,"edl_appt_prix_f1",client.id)
                        checktarif(ad,"edl_appt_prix_f2",client.id)
                        checktarif(ae,"edl_appt_prix_f3",client.id)
                        checktarif(af,"edl_appt_prix_f4",client.id)
                        checktarif(ag,"edl_appt_prix_f5",client.id)
                        checktarif(ah,"edl_appt_prix_f6",client.id)
                        checktarif(ai,"edl_pav_villa_prix_t1",client.id)
                        checktarif(aj,"edl_pav_villa_prix_t2",client.id)
                        checktarif(ak,"edl_pav_villa_prix_t3",client.id)
                        checktarif(al,"edl_pav_villa_prix_t4",client.id)
                        checktarif(am,"edl_pav_villa_prix_t5",client.id)
                        checktarif(an,"edl_pav_villa_prix_t6",client.id)
                        checktarif(ao,"edl_pav_villa_prix_t7",client.id)
                        checktarif(ap,"edl_pav_villa_prix_t8",client.id)
                        checktarif(aq,"chif_appt_prix_stu",client.id)
                        checktarif(ar,"chif_appt_prix_f1",client.id)
                        checktarif(As,"chif_appt_prix_f2",client.id)
                        checktarif(at,"chif_appt_prix_f3",client.id)
                        checktarif(au,"chif_appt_prix_f4",client.id)
                        checktarif(av,"chif_appt_prix_f5",client.id)
                        checktarif(aw,"chif_pav_villa_prix_t1",client.id)
                        checktarif(ax,"chif_pav_villa_prix_t2",client.id)
                        checktarif(ay,"chif_pav_villa_prix_t3",client.id)
                        checktarif(az,"chif_pav_villa_prix_t4",client.id)
                        checktarif(ba,"chif_pav_villa_prix_t5",client.id)
                        checktarif(bc,"chif_pav_villa_prix_t6",client.id)
                        checktarif(bd,"chif_pav_villa_prix_t7",client.id)
                        checktarif(be,"chif_pav_villa_prix_t8",client.id)

    '''for (m,n,o,p,q,r,s,t,u,v,x,y,z,aa,bb,cc,dd,ee,ff,ab,ac,ad,ae,af,ag,ah,ai,aj,ak,al,am,an,ao,ap,aq,ar,As,at,au,av,aw,ax,ay,az,ba,bc,bd,be,bf,ca,km,ck) in zip(df['AS du Client'],df['% Com AS du client'],df['Nom Respon Cell Dev'],
                                                        df['% Respon Cell Dev'],df['Nom agent Cell Dev'],df['% Agent Cell Dev']
                                                        ,df['Nom Agent TECH EDELE'],df['% Agent TECH EDELE'],df['Nom Respon Cell Tech']
                                                        ,df['% Respon Cell Tech'],df['Nom Suiveur Cell Tech'],df['% Suiveur Cell Tech'],
                                                        df['Nom Respon Cell Planif'],df['% Respon Cell Planif'],df['Nom Suiveur Cell Planif'],
                                                        df['% Suiveur Cell Planif'],df['Nom Agent saisie Cell Planif'],df['% Agent saisie CEll planif'],
                                                        df['TAUX DE MAJORATION MEUBLE'],df['EDL_PRIX STD'],df['EDL_APPT / PRIX T1'],df['EDL_APPT / PRIX T2'],
                                                        df['EDL_APPT / PRIX T3'],df['EDL_APPT / PRIX T4'],df['EDL_APPT / PRIX T5'],df['EDL_APPT / PRIX T6'],
                                                        df['EDL_PAV VILLA / PRIX T1'],df['EDL_PAV VILLA / PRIX T2'],df['EDL_PAV VILLA / PRIX T3'],
                                                        df['EDL_PAV VILLA / PRIX T4'],df['EDL_PAV VILLA / PRIX T5'],df['EDL_PAV VILLA / PRIX T6'],
                                                        df['EDL_PAV VILLA / PRIX T7'],df['EDL_PAV VILLA / PRIX T8'],df['CHIF_APPT / PRIX STU'],
                                                        df['CHIF_APPT / PRIX T1'],df['CHIF_APPT / PRIX T2'],df['CHIF_APPT / PRIX T3'],
                                                        df['CHIF_APPT / PRIX T4'],df['CHIF_APPT / PRIX T5'],df['CHIF_PAV VILLA / PRIX T1'],
                                                        df['CHIF_PAV VILLA / PRIX T2'],df['CHIF_PAV VILLA / PRIX T3'],df['CHIF_PAV VILLA / PRIX T4'],
                                                        df['CHIF_PAV VILLA / PRIX T5'],df['CHIF_PAV VILLA / PRIX T6'],df['CHIF_PAV VILLA / PRIX T7'],
                                                        df['CHIF_PAV VILLA / PRIX T8'],df['PRIX AUTRE'],df['Ref proposition commerciale']
                                                        ,df['Référence de la société'],df['Nom du ou des responsables']):
                        try:
                            cli=Client.query.filter(and_(Client.reference==ca,Client.societe==km,Client.nom==str(ck.lower()))).first() #check by reference number,societekm, and nomck
                        except:
                            cli=Client.query.filter(and_(Client.reference==ca,Client.societe==km)).first() #check by reference number,societekm, and nomck
                        tarif =Tarifs.query.filter_by(reference_client=int(cli.id)).first()
                        try:
                            r_C=Expert.query.filter_by(nom=str(m.lower())).first()
                            if r_C is not None:#check keys
                                rC= r_C.id
                            if r_C is None  :
                                rC= 0
                        except:
                            rC= 0
                        try:
                            r_r=Expert.query.filter_by(nom=str(o.lower())).first()
                            if r_r is not None:
                                rr= r_r.id
                            if r_r is None  :
                                rr= 0
                        except:
                            rr=0
                        try:
                            dr_a=Expert.query.filter_by(nom=str(q.lower())).first()
                            if dr_a is not None:
                                dra= dr_a.id
                            if dr_a is None  :
                                dra= 0
                        except:
                            dra= 0
                        try:
                            tr_a=Expert.query.filter_by(nom=str(s.lower())).first()
                            if tr_a is not None:
                                tra= tr_a.id
                            if tr_a is None  :
                                tra= 0
                        except:
                            tra=0
                        try:
                            tr_r=Expert.query.filter_by(nom=str(u.lower())).first()
                            if tr_r is not None:
                                trr= tr_r.id
                            if tr_r is None  :
                                trr= 0
                        except:
                            trr= 0
                        try:
                            tr_s=Expert.query.filter_by(nom=str(x.lower())).first()
                            if tr_s is not None:
                                trs= tr_s.id
                            if tr_s is None  :
                                trs= 0
                        except:
                            trs= 0
                        try:
                            pr_s=Expert.query.filter_by(nom=str(z.lower())).first()
                            if pr_s is not None:
                                prs= pr_s.id
                            if pr_s is None  :
                                prs= 0
                        except:
                            prs=0
                        try:
                            pr_si=Expert.query.filter_by(nom=str(bb.lower())).first()
                            if pr_si is not None:
                                prsi= pr_si.id
                            if pr_si is None  :
                                prsi= 0
                        except:
                            prsi= 0
                        try:
                            ra_s=Expert.query.filter_by(nom=str(dd.lower())).first()
                            if ra_s is not None:
                                ras= ra_s.id
                            if ra_s is None  :
                                ras= 0  
                        except:
                            ras= 0    
                        if tarif is None:
                            taf_base =Tarifs(reference_client=cli.id,
                            referent_as_client=rC,com_as_sur_ca_client=str(n),cell_dev_ref_responsable=rr,
                            com_cell_dev_ref_responsable=str(p),cell_dev_ref_agent=dra,com_cell_dev_ref_agent=str(r),
                            cell_tech_ref_agent=tra,com_cell_tech_Ref_agent=str(t),cell_tech_ref_responsable=trr,
                            com_cell_tech_ref_responsable=str(v),cell_tech_ref_suiveur=trs,
                            com_cell_tech_ref_suiveur=str(y),cell_planif_ref_responsable=prs,
                            com_cell_planif_ref_responsable=str(aa),cell_planif_ref_suiveur=prsi,
                            com_cell_planif_ref_suiveur=str(cc),cell_planif_ref_agent_saisie=ras,
                            com_cell_planif_ref_agent_saisie=str(ee),taux_meuble=str(ff),prix_autre=str(bf))
                            
                            db.session.add(taf_base)
                            db.session.commit()
                            checktarif(ab,taf_base.edl_prix_std)
                            checktarif(ac,edl_appt_prix_f1)
                            checktarif(ad,edl_appt_prix_f2)
                            checktarif(ae,edl_appt_prix_f3)
                            checktarif(af,edl_appt_prix_f4)
                            checktarif(ag,edl_appt_prix_f5)
                            checktarif(ah,edl_appt_prix_f6)
                            checktarif(ai,edl_pav_villa_prix_t1)
                            checktarif(aj,edl_pav_villa_prix_t2)
                            checktarif(ak,edl_pav_villa_prix_t3)
                            checktarif(al,edl_pav_villa_prix_t4)
                            checktarif(am,edl_pav_villa_prix_t5)
                            checktarif(an,edl_pav_villa_prix_t6)
                            checktarif(ao,edl_pav_villa_prix_t7)
                            checktarif(ap,edl_pav_villa_prix_t8)
                            checktarif(aq,chif_appt_prix_stu)
                            checktarif(ar,chif_appt_prix_f1)
                            checktarif(As,chif_appt_prix_f2)
                            checktarif(at,chif_appt_prix_f3)
                            checktarif(au,chif_appt_prix_f4)
                            checktarif(av,chif_appt_prix_f5)
                            checktarif(aw,chif_pav_villa_prix_t1)
                            checktarif(ax,chif_pav_villa_prix_t2)
                            checktarif(ay,chif_pav_villa_prix_t3)
                            checktarif(az,chif_pav_villa_prix_t4)
                            checktarif(ba,chif_pav_villa_prix_t5)
                            checktarif(bc,chif_pav_villa_prix_t6)
                            checktarif(bd,chif_pav_villa_prix_t7)
                            checktarif(be,chif_pav_villa_prix_t8)
                             
                        else:
                            print('exist2')'''

def expert__(loc):
    
    df = pd.read_excel(loc)
    for (ca,a,b,c,d,e,f,g,h,i,j,k,l,m,n,o,p,q,r,s,t,u,v,y,z,aa,bb,cc) in zip(df['ID'],df['identité agent'],df['Téléphone'],df['AS AC AP'],
                                                                            df['email groupe'],df['email perso'],df['Code Tva'],
                                                                            df['Taux Tva'],df['SIRET'],df['date entrée'],df['Trigramme'],
                                                                            df['secteur activité'],df['Adresse 1'],df['Adresse 2'],df['CP'],
                                                                            df['Ville'],df['Login backof'],df['PWD backof.1'],df['Login extranet']
                                                                            ,df['Pwd extranet'],df['Login tablette'],df['PWD backof'],df['Observations de suivi'],
                                                                            df['Actif  Parti'],df['Type certification'],df['Date certification initiale']
                                                                            ,df['date renouv certification'],df['Pwd Gsuite']):
            cli=Expert.query.filter_by(nom=str(a.lower())).first()

            if cli is None:
                expert=Expert(nom=a.lower(),numero=b,TYPE=c,email=d,email_perso=e,code_tva=f,taux_tva=g,siret=h,date_entree=i,
                trigramme=j)    
                db.session.add(expert)
                db.session.commit()
                history=Expert_History(expert_id=expert.id,secteur=k,adresse1=l,adresse2=m,cp=n,ville=o,login_backof=p,pwd_backof=q
                ,login_extranet=r,pwd_extranet=s,login_google=t,pwd_google=u,observations_de_suivi=v,actif_parti=y,type_certification=z,
                date_certification_initiale=aa,date_renouv_certification=bb,pwd_gsuite=cc)
                db.session.add(history)
                db.session.commit()
            else:
                print('already exist')

def Missions(loc):
    
    print('ok')
    wb_obj = openpyxl.load_workbook(loc,data_only=True)
    print(wb_obj.sheetnames)
    sheet=wb_obj['Sheet1']
    print(sheet)
    reference=[]
    Nom=[]
    num=[]
    
    for i in range(1,sheet.max_row):
        
        print(sheet["I"][i].value)
        cli=Client.query.filter_by(reference=str(sheet["A"][i].value)).first()
        if cli is None:
            cli=Client.query.filter_by(nom=str(sheet["D"][i].value.lower())).first()
        '''except:
            cli=Client.query.filter_by(nom=str(sheet["B"][i].value.lower())).first()
        if cli1 is None:
            try:
                client=Client(reference=str(sheet["A"][i].value),nom=str(sheet["B"][i].value.lower()))#,date_creation=e
            except:
                client=Client(reference=str(sheet["A"][i].value))#,date_creation=e
            db.session.add(client)
            db.session.commit()   
            history=Client_History(client_id=client.id)
            db.session.commit() '''
        
        #try:
          #  cli=Client.query.filter_by(reference=str(sheet["A"][i].value)).first()
        #except:
           # cli1=Client.query.filter_by(nom=str(sheet["B"][i].value)).first()
        if sheet["L"][i].value is None:
            SA=0
        else:
            AS=Expert.query.filter_by(nom=str(sheet["L"][i].value.lower())).first()
            if AS is not None:
                SA= AS.id
            if AS is None :
                SA= 0
        if sheet["R"][i].value is None:
            IV=0
        else:
            INTERV=Expert.query.filter_by(nom=str(sheet["R"][i].value.lower())).first()
            if INTERV is not None:
                IV= INTERV.id
            if INTERV is None  :
                IV= 0
        if sheet["AL"][i].value is None:
            MC=0
        else:
            M_C=Expert.query.filter_by(nom=str(sheet["AL"][i].value.lower())).first()
            if M_C is not None:
                MC= M_C.id
            if M_C is None  :
                MC= 0
        if sheet["AN"][i].value is None:
            AC=0
        else:
            A_C=Expert.query.filter_by(nom=str(sheet["AN"][i].value.lower())).first()
            if A_C is not None:
                AC= A_C.id
            if A_C is None  :
                AC= 0

        if sheet["BR"][i].value is None:
            RCD=0
        else:
            R_CD=Expert.query.filter_by(nom=str(sheet["BR"][i].value.lower())).first()
            if R_CD is not None:
                RCD= R_CD.id
            if R_CD is None  :
                RCD= 0

        if sheet["BT"][i].value is None:
            ACD=0
        else:
            A_CD=Expert.query.filter_by(nom=str(sheet["BT"][i].value.lower())).first()
            if A_CD is not None:
                ACD= A_CD.id
            if A_CD is None  :
                ACD= 0

        if sheet["BV"][i].value is None:
            ACT=0
        else:
            A_CT=Expert.query.filter_by(nom=str(sheet["BV"][i].value.lower())).first()
            if A_CT is not None:
                ACT= A_CT.id
            if A_CT is None  :
                ACT= 0

        if sheet["BX"][i].value is None:
            RCT=0
        else:
            R_CT=Expert.query.filter_by(nom=str(sheet["BX"][i].value.lower())).first()
            if R_CT is not None:
                RCT= R_CT.id
            if R_CT is None  :
                RCT= 0

        if sheet["BZ"][i].value is None:
            SCT=0
        else:
            S_CT=Expert.query.filter_by(nom=str(sheet["BZ"][i].value.lower())).first()
            if S_CT is not None:
                SCT= S_CT.id
            if S_CT is None  :
                SCT= 0

        if sheet["CB"][i].value is None:
            RCP=0
        else:
            R_CP=Expert.query.filter_by(nom=str(sheet["CB"][i].value.lower())).first()
            if R_CP is not None:
                RCP= R_CP.id
            if R_CP is None  :
                RCP= 0

        if sheet["CD"][i].value is None:
            SCP=0
        else:
            S_CP=Expert.query.filter_by(nom=str(sheet["CD"][i].value.lower())).first()
            if S_CP is not None:
                SCP= S_CP.id
            if S_CP is None  :
                SCP= 0

        if sheet["CF"][i].value is None:
            ASCP=0
        else:
            AS_CP=Expert.query.filter_by(nom=str(sheet["CF"][i].value.lower())).first()
            if AS_CP is not None:
                ASCP= AS_CP.id
            if AS_CP is None  :
                ASCP= 0

        if cli:
            mission=Mission(Reference_BAILLEUR=cli.id,
           # old=sheet["A"][i].value,
            ABONNEMENT	 = sheet["J"][i].value ,
            ID_AS	 = SA ,
        
          
            DATE_REALISE_EDL =sheet["M"][i].value , 	
            ID_INTERV = IV ,
            NRO_FACTURE = sheet["I"][i].value ,
            PRIX_HT_EDL = sheet["N"][i].value ,
            TVA_EDL = sheet["O"][i].value ,
            PRIX_TTC_EDL = sheet["P"][i].value ,
            Reference_LOCATAIRE	 =  sheet["U"][i].value ,
            Adresse1_Bien	 = sheet["V"][i].value ,  
            Adresse2_Bien	 = sheet["W"][i].value , 
            CP_Bien	 = sheet["X"][i].value ,  
            Ville_Bien	 = sheet["Y"][i].value , 
            
            CA_HT_AS = sheet["Z"][i].value , 	
            TVA_AS	 = sheet["AA"][i].value , 
            CA_TTC_AS = sheet["AB"][i].value , 	
            CA_HT_AC = sheet["AC"][i].value , 	
            CA_TTC_AC	 = sheet["AD"][i].value , 
            CA_HT_TRUST	 = sheet["AE"][i].value , 
            TVA_TRUST	 = sheet["AF"][i].value ,
            Prix_ht_chiffrage	 = sheet["AH"][i].value , 
            POURCENTAGE_suiveur_chiffrage	 = sheet["AI"][i].value ,
            POURCENTAGE_AS_chiffrage = sheet["AJ"][i].value ,	
            POURCENTAGE_manager_chiffrage  = sheet["AK"][i].value , 	
            ID_manager_chiffrage  = MC ,
                
            POURCENTAGE_agent_chiffrage = sheet["AM"][i].value ,	
            ID_agent_chiffrage  = AC ,	
            
            TYPE_EDL = sheet["AO"][i].value ,	
            TITREPROPRIO = sheet["AQ"][i].value , 		
            NOMPROPRIO = sheet["AR"][i].value , 	
            DATE_FACT_REGLEE = sheet["AS"][i].value ,	
            TYPE_LOGEMENT = sheet["BB"][i].value , 	
            CODE_AMEXPERT = sheet["AI"][i].value , 	
            COMMENTAIRE_FACTURE = sheet["BC"][i].value , 	
            LOGEMENT_MEUBLE =sheet["BH"][i].value , 	
            CODE_FACTURATION = sheet["BI"][i].value , 	
            TYPE_DE_BIEN = sheet["BJ"][i].value , 	
            surface_logement1 = sheet["BK"][i].value , 		
            DATE_FACTURE = sheet["AP"][i].value , 	
            POURCENTAGE_COM_AS_DU_CLIENT = sheet["BQ"][i].value , 	
            ID_Respon_Cell_Dev	 =RCD ,
            
            POURCENTAGE_Respon_Cell_Dev = sheet["BS"][i].value , 	
            ID_agent_Cell_Dev = ACD, 	
            
            POURCENTAGE_Agent_Cell_Dev = sheet["BU"][i].value ,	
            ID_Agent_CellTech = ACT,  	
            
            POURCENTAGE_Agent_Cell_Tech = sheet["BW"][i].value , 	
            ID_Respon_Cell_Tech = RCT, #######
                
            POURCENTAGE_Respon_Cell_Tech = sheet["BY"][i].value ,	
            ID_Suiveur_Cell_Tech  = SCT ,
            
            POURCENTAGE_Suiveur_Cell_Tech	 = sheet["CA"][i].value , 
            ID_Respon_Cell_Planif = RCP,
            
            POURCENTAGE_Respon_Cell_Planif  = sheet["CC"][i].value ,
            ID_Suiveur_Cell_Planif  = SCP,
            
            POURCENTAGE_Suiveur_Cell_Planif	 = sheet["CE"][i].value , 
            ID_Agent_saisie_Cell_Planif  = ASCP,
                  
            POURCENTAGE_Agent_saisie_CEll_planif  = sheet["CG"][i].value )
            db.session.add(mission)
            db.session.commit()
            
        else:

            Nom.append(sheet["D"][i].value.lower())
            num.append(i)
            reference.append((sheet["A"][i].value))# make a table for historique des donnees 
            print(reference)
            print(Nom)
            print(num)
        

def fix_mission():
    mission=Mission.query.all()
    for i in mission:
        try:
            if i.CODE_FACTURATION[-1]=='M':
                print(i.CODE_FACTURATION)
                i.CODE_FACTURATION[2:5] == 150
                A=i.CODE_FACTURATION[0:-1]
                i.CODE_FACTURATION = A
                print(i.CODE_FACTURATION)
                db.session.commit()

            if i.CODE_FACTURATION[-1] == ' ':
                print( i.CODE_FACTURATION)
                B=i.CODE_FACTURATION[0:-1] 
                i.CODE_FACTURATION = B
                db.session.commit()

            if i.CODE_FACTURATION[-2:] == "00" or  i.CODE_FACTURATION[-2:] == "50" :
                    i.Anomalie = False
                    i.reason = None
                    db.session.commit()
        except:
            i.coherent = False
            i.reason = "Anomalie bloquante codification du code facturation incorrect   "
            print(i)
            db.session.commit()
        
        try:
            if i.TYPE_LOGEMENT[-1] == ' ':
                print( i.TYPE_LOGEMENT)
                B=i.TYPE_LOGEMENT[0:-1] 
                i.TYPE_LOGEMENT = B
                db.session.commit()


            if i.TYPE_LOGEMENT[-1] == 'M':
                print(i.TYPE_LOGEMENT)
                i.CODE_FACTURATION[2:5] == 150
                print(i.CODE_FACTURATION)
                B=i.TYPE_LOGEMENT[0:-1] 
                i.TYPE_LOGEMENT = B
                print(i.TYPE_LOGEMENT)
                db.session.commit()

            if i.TYPE_LOGEMENT[-2:] != i.CODE_FACTURATION[-2:]  :
                i.Anomalie = True
                i.reason = "Anomalie non bloquante traite en  "+str(i.TYPE_LOGEMENT[-2:])
                db.session.commit()

            

            if len(i.TYPE_LOGEMENT[0:4]) < 3 :
                if len(i.TYPE_LOGEMENT[0:2]) == 2:
                    if i.TYPE_LOGEMENT[0] == "T":
                        B="PAV-"+str(i.TYPE_LOGEMENT) 
                        i.TYPE_LOGEMENT = B
                        db.session.commit()
                    if i.TYPE_LOGEMENT[0] == "F":
                        B="APPT-"+str(i.TYPE_LOGEMENT)
                        i.TYPE_LOGEMENT = B
                        db.session.commit() 
                else:
                    i.coherent = False 
                    i.reason = "Anomalie bloquante codification du type de logement incorrect sur  "+str(i.TYPE_LOGEMENT)
                    db.session.commit()
            
        except:
            i.coherent = False
            i.reason = "Anomalie bloquante codification du type de logement incorrect "
            db.session.commit()
        


def reset():
    mission=Mission.query.delete()
    db.session.commit()
    #for i in mission:
        #i.NRO_FACTURE = None
       # db.session.commit()[None]
#['bertay fabrice']
#[1362]
#[None, 120252]
##['bertay fabrice', 'detoc olivier et latitia']
#[1362, 1365]
#[None, 120252, 108746]
#['bertay fabrice', 'detoc olivier et latitia', 'boivin - champeaux']
#[1362, 1365, 1836]
       # print('o')Histo4723l_22091-22556

def Base(loc):
    wb_obj = openpyxl.load_workbook(loc)
    sheet=wb_obj.active
    A=['F','G','H','I','J','K','L']

    for i in A:
        base_appt=Tarif_base(pav_appartement='APPT',Type=sheet[i][24].value,surface=sheet[i][25].value,Prix_EDL=sheet[i][26].value)
        base_pav=Tarif_base(pav_appartement='PAV',Type=sheet[i][28].value,surface=sheet[i][29].value,Prix_EDL=sheet[i][30].value)
        if sheet[i][28].value == 'T2':
            base_pav1=Tarif_base(pav_appartement='PAV',Type='T1',surface=0,Prix_EDL=0)
            db.session.add(base_pav1)
            db.session.commit()
        db.session.add(base_appt)
        db.session.add(base_pav)
        db.session.commit()
        


    
#def suivi(loc):




def Missions2(loc,n):
    
    wb_obj = openpyxl.load_workbook(loc,data_only=True)
    print(wb_obj.sheetnames)
    sheet=wb_obj[n]
    
    reference=[]
    
    for i in range(1,sheet.max_row):
        a=str(sheet["B"][i].value.lower()).split()
        v=''.join(a)
        vii=v.split("-")
        vii=''.join(vii)
        cli=Client.query.filter_by(enseigne=str(vii)).first()
        if cli:
            mission=Mission(Reference_BAILLEUR=cli.id,
           # old=sheet["A"][i].value,
            ABONNEMENT	 = sheet["J"][i].value ,
            ID_AS	 = 0 ,
        
          
            DATE_REALISE_EDL =sheet["M"][i].value , 	
            ID_INTERV = 0 ,
            NRO_FACTURE = sheet["I"][i].value ,
            PRIX_HT_EDL = sheet["N"][i].value ,
            TVA_EDL = sheet["O"][i].value ,
            PRIX_TTC_EDL = sheet["P"][i].value ,
            Reference_LOCATAIRE	 =  sheet["U"][i].value ,
            Adresse1_Bien	 = sheet["V"][i].value ,  
            Adresse2_Bien	 = sheet["W"][i].value , 
            CP_Bien	 = sheet["X"][i].value ,  
            Ville_Bien	 = sheet["Y"][i].value , 
            
            CA_HT_AS = sheet["Z"][i].value , 	
            TVA_AS	 = sheet["AA"][i].value , 
            CA_TTC_AS = sheet["AB"][i].value , 	
            CA_HT_AC = sheet["AC"][i].value , 	
            CA_TTC_AC	 = sheet["AD"][i].value , 
            CA_HT_TRUST	 = sheet["AE"][i].value , 
            TVA_TRUST	 = sheet["AF"][i].value ,
            Prix_ht_chiffrage	 = sheet["AH"][i].value , 
            POURCENTAGE_suiveur_chiffrage	 = sheet["AI"][i].value ,
            POURCENTAGE_AS_chiffrage = sheet["AJ"][i].value ,	
            POURCENTAGE_manager_chiffrage  = sheet["AK"][i].value , 	
            ID_manager_chiffrage  = 0 ,
                
            POURCENTAGE_agent_chiffrage = sheet["AM"][i].value ,	
            ID_agent_chiffrage  = 0 ,	
            
            TYPE_EDL = sheet["AO"][i].value ,	
            TITREPROPRIO = sheet["AQ"][i].value , 		
            NOMPROPRIO = sheet["AR"][i].value , 	
            DATE_FACT_REGLEE = sheet["AS"][i].value ,	
            TYPE_LOGEMENT = sheet["AX"][i].value , 	
            CODE_AMEXPERT = sheet["BB"][i].value , 	
            COMMENTAIRE_FACTURE = sheet["BC"][i].value , 	
            LOGEMENT_MEUBLE =sheet["BH"][i].value , 	
            CODE_FACTURATION = sheet["BI"][i].value , 	
            TYPE_DE_BIEN = sheet["BJ"][i].value , 	
            surface_logement1 = sheet["BK"][i].value , 		
            DATE_FACTURE = sheet["I"][i].value , 	
            POURCENTAGE_COM_AS_DU_CLIENT = sheet["BQ"][i].value , 	
            ID_Respon_Cell_Dev	 =0 ,
            
           
            ID_agent_Cell_Dev = 0, 	
            
           
            ID_Agent_CellTech = 0,  	
            
            	
            ID_Respon_Cell_Tech = 0, #######
                
            	
            ID_Suiveur_Cell_Tech  = 0 ,
            
            
            ID_Respon_Cell_Planif = 0,
            
           
            ID_Suiveur_Cell_Planif  = 0,
            
             
            ID_Agent_saisie_Cell_Planif  = 0
                  
             )
            db.session.add(mission)
            db.session.commit()
            try:
                if mission.CODE_FACTURATION[-1]=='M':
                    print(mission.CODE_FACTURATION[2:5])
                    mission.CODE_FACTURATION[2:5] == 150
                    A=mission.CODE_FACTURATION[0:-1]
                    mission.CODE_FACTURATION = A
                    #print(mission.CODE_FACTURATION)
                    db.session.commit()

                if mission.CODE_FACTURATION[-1] == ' ':
                    print(mission.CODE_FACTURATION[2:5])
                    B=mission.CODE_FACTURATION[0:-1] 
                    mission.CODE_FACTURATION = B
                    db.session.commit()

                if mission.CODE_FACTURATION[-2:] == "00" or  mission.CODE_FACTURATION[-2:] == "50" :
                        mission.Anomalie = False
                        mission.reason = None
                        db.session.commit()
            except:
                mission.coherent = False
                mission.reason = "Anomalie bloquante codification du code facturation incorrect sur  "
               # print(mission)
                db.session.commit()
            print(mission.CODE_FACTURATION)
            print(mission.TYPE_LOGEMENT)
            try:
                if mission.TYPE_LOGEMENT[-1] == ' ':
                   # print( mission.TYPE_LOGEMENT)
                    B=mission.TYPE_LOGEMENT[0:-1] 
                    mission.TYPE_LOGEMENT = B
                    db.session.commit()


                if mission.TYPE_LOGEMENT[-1] == 'M':
                   # print(mission.TYPE_LOGEMENT)
                    mission.CODE_FACTURATION[2:5] == 150
                    #print(mission.CODE_FACTURATION)
                    B=mission.TYPE_LOGEMENT[0:-1] 
                    mission.TYPE_LOGEMENT = B
                   # print(mission.TYPE_LOGEMENT)
                    db.session.commit()

                if mission.TYPE_LOGEMENT[-2:] != mission.CODE_FACTURATION[-2:]  :
                    mission.Anomalie = True
                    mission.reason = "Anomalie non bloquante traite en  "+str(mission.TYPE_LOGEMENT[-2:])
                    db.session.commit()

                

                if len(mission.TYPE_LOGEMENT[0:4]) < 3 :
                    if len(mission.TYPE_LOGEMENT[0:2]) == 2:
                        if mission.TYPE_LOGEMENT[0] == "T":
                            B="PAV-"+str(mission.TYPE_LOGEMENT) 
                            mission.TYPE_LOGEMENT = B
                            db.session.commit()
                        if mission.TYPE_LOGEMENT[0] == "F":
                            B="APPT-"+str(mission.TYPE_LOGEMENT)
                            mission.TYPE_LOGEMENT = B
                            db.session.commit() 
                    else:
                        mission.coherent = False 
                        mission.reason = "Anomalie bloquante codification du type de logement incorrect sur  "+str(mission.TYPE_LOGEMENT)
                        db.session.commit()
                
            except:
                mission.coherent = False
                mission.reason = "Anomalie bloquante codification du type de logement incorrect "
                db.session.commit()
            try:
                if mission.TYPE_LOGEMENT[-1] == ' ':
                   # print( mission.TYPE_LOGEMENT)
                    B=mission.TYPE_LOGEMENT[0:-1] 
                    mission.TYPE_LOGEMENT = B
                    db.session.commit()


                if mission.TYPE_LOGEMENT[-1] == 'M':
                   # print(mission.TYPE_LOGEMENT)
                    mission.CODE_FACTURATION[2:5] == 150
                    #print(mission.CODE_FACTURATION)
                    B=mission.TYPE_LOGEMENT[0:-1] 
                    mission.TYPE_LOGEMENT = B
                   # print(mission.TYPE_LOGEMENT)
                    db.session.commit()

                if mission.TYPE_LOGEMENT[-2:] != mission.CODE_FACTURATION[-2:]  :
                    mission.Anomalie = True
                    mission.reason = "Anomalie non bloquante traite en  "+str(mission.TYPE_LOGEMENT[-2:])
                    db.session.commit()

                

                if len(mission.TYPE_LOGEMENT[0:4]) < 3 :
                    if len(mission.TYPE_LOGEMENT[0:2]) == 2:
                        if mission.TYPE_LOGEMENT[0] == "T":
                            B="PAV-"+str(mission.TYPE_LOGEMENT) 
                            mission.TYPE_LOGEMENT = B
                            db.session.commit()
                        if mission.TYPE_LOGEMENT[0] == "F":
                            B="APPT-"+str(mission.TYPE_LOGEMENT)
                            mission.TYPE_LOGEMENT = B
                            db.session.commit() 
                    else:
                        mission.coherent = False 
                        mission.reason = "Anomalie bloquante codification du type de logement incorrect sur  "+str(mission.TYPE_LOGEMENT)
                        db.session.commit()
                
            except:
                mission.coherent = False
                mission.reason = "Anomalie bloquante codification du type de logement incorrect "
                db.session.commit()
            
            if mission.TYPE_LOGEMENT == None:
                mission.TYPE_LOGEMENT = ''
                db.session.commit()
            if mission.CODE_FACTURATION == None:
                mission.CODE_FACTURATION = ''
                db.session.commit()
            try:
                if mission.CODE_FACTURATION[0:2] == 'TS':
                    mission.PRIX_HT_EDL = mission.CODE_FACTURATION[2:5]
                    db.session.commit()
                if mission.CODE_FACTURATION[0:2] == 'TN':
                
                    if mission.TYPE_LOGEMENT[0:4] == 'APPT' and mission.TYPE_LOGEMENT[-1] == '1'   :
                        
                        tarif=Tarifs.query.filter_by(reference_client = mission.Reference_BAILLEUR).first()
                        if mission.CODE_FACTURATION[2:5] == '150':#check print# fix
                            meuble=float(tarif.edl_appt_prix_f1)/2
                            mission.PRIX_HT_EDL = float(tarif.edl_appt_prix_f1) + float(meuble)
                            db.session.commit()
                            
                            
                        else:
                            
                            mission.PRIX_HT_EDL = tarif.edl_appt_prix_f1
                            db.session.commit()
                    if mission.TYPE_LOGEMENT[0:4] == 'APPT' and mission.TYPE_LOGEMENT[-1] == 'U'   :
                        
                        tarif=Tarifs.query.filter_by(reference_client = mission.Reference_BAILLEUR).first()
                        if mission.CODE_FACTURATION[2:5] == '150':#check print# fix
                            meuble=float(tarif.edl_prix_std)/2
                            mission.PRIX_HT_EDL = float(tarif.edl_prix_std) + float(meuble)
                            db.session.commit()
                            
                            
                        else:
                            
                            mission.PRIX_HT_EDL = tarif.edl_appt_prix_f1
                            db.session.commit()

                    if mission.TYPE_LOGEMENT[0:4] == 'APPT' and mission.TYPE_LOGEMENT[-1] == '2':
                        #print('APPT 2')
                        tarif=Tarifs.query.filter_by(reference_client = mission.Bailleur__data.id).first()
                        if mission.CODE_FACTURATION[2:5] == '150':#check print
                            meuble=float(tarif.edl_appt_prix_f2)/2
                            mission.PRIX_HT_EDL = float(tarif.edl_appt_prix_f2) + float(meuble)
                            db.session.commit()
                            
                        else:
                            mission.PRIX_HT_EDL = tarif.edl_appt_prix_f2
                            db.session.commit()
                            
                    if mission.TYPE_LOGEMENT[0:4] == 'APPT' and mission.TYPE_LOGEMENT[-1] == '3':
                        tarif=Tarifs.query.filter_by(reference_client = mission.Bailleur__data.id).first()
                        if mission.CODE_FACTURATION[2:5] == '150':#check print
                            meuble=float(tarif.edl_appt_prix_f3)/2
                            mission.PRIX_HT_EDL = float(tarif.edl_appt_prix_f3) + float(meuble)
                            db.session.commit()
                            
                        else:
                            mission.PRIX_HT_EDL = tarif.edl_appt_prix_f3
                            db.session.commit()
                            
                    if mission.TYPE_LOGEMENT[0:4] == 'APPT' and mission.TYPE_LOGEMENT[-1] == '4':
                        
                        tarif=Tarifs.query.filter_by(reference_client = mission.Bailleur__data.id).first()
                        if mission.CODE_FACTURATION[2:5] == '150':#check print
                            meuble=float(tarif.edl_appt_prix_f4)/2
                            mission.PRIX_HT_EDL = float(tarif.edl_appt_prix_f4) + float(meuble)
                            db.session.commit()
                            
                        else:
                            mission.PRIX_HT_EDL = tarif.edl_appt_prix_f4
                            db.session.commit()
                            
                    if mission.TYPE_LOGEMENT[0:4] == 'APPT' and mission.TYPE_LOGEMENT[-1] == '5':
                        tarif=Tarifs.query.filter_by(reference_client = mission.Bailleur__data.id).first()
                        if mission.CODE_FACTURATION[2:5] == '150':#check print
                            meuble=float(tarif.edl_appt_prix_f5)/2
                            mission.PRIX_HT_EDL = float(tarif.edl_appt_prix_f5) + float(meuble)
                            db.session.commit()
                            
                        else:
                            mission.PRIX_HT_EDL = tarif.edl_appt_prix_f5
                            db.session.commit()
                            
                    if mission.TYPE_LOGEMENT[0:4] == 'APPT' and mission.TYPE_LOGEMENT[-1] == '6':
                        tarif=Tarifs.query.filter_by(reference_client = mission.Bailleur__data.id).first()
                        if mission.CODE_FACTURATION[2:5] == '150':#check print
                            meuble=float(tarif.edl_appt_prix_f6)/2
                            mission.PRIX_HT_EDL = float(tarif.edl_appt_prix_f6) + float(meuble)
                            db.session.commit()
                            
                        else:
                            mission.PRIX_HT_EDL = tarif.edl_appt_prix_f6
                            db.session.commit()
                            
                    if mission.TYPE_LOGEMENT[0:4] == 'APPT' and mission.TYPE_LOGEMENT[-1] == '7':
                        tarif=Tarifs.query.filter_by(reference_client = mission.Bailleur__data.id).first()
                        if mission.CODE_FACTURATION[2:5] == '150':#check print
                            meuble=float(tarif.EDL_APPT_prix_F7)/2
                            mission.PRIX_HT_EDL = float(tarif.edl_appt_prix_f7) + float(meuble)
                            db.session.commit()
                            
                        else:
                            mission.PRIX_HT_EDL = tarif.edl_appt_prix_f7
                            db.session.commit()
                            

                    if mission.TYPE_LOGEMENT[0:3] == 'PAV' and mission.TYPE_LOGEMENT[-1] == '1' :
                        tarif=Tarifs.query.filter_by(reference_client=mission.Bailleur__data.id).first()
                        if mission.CODE_FACTURATION[2:5] == '150':
                            meuble=float(tarif.edl_pav_villa_prix_t1)/2
                            mission.PRIX_HT_EDL = float(tarif.edl_pav_villa_prix_t1) + float(meuble)
                            db.session.commit()
                            
                        else:
                            mission.PRIX_HT_EDL = tarif.edl_pav_villa_prix_t1
                            db.session.commit()
                            
                    if mission.TYPE_LOGEMENT[0:3] == 'PAV' and mission.TYPE_LOGEMENT[-1] == '2' :
                        
                        tarif=Tarifs.query.filter_by(reference_client=mission.Bailleur__data.id).first()
                        if mission.CODE_FACTURATION[2:5] == '150':
                            meuble=float(tarif.edl_pav_villa_prix_t2)/2
                            mission.PRIX_HT_EDL = float(tarif.edl_pav_villa_prix_t2) + float(meuble)
                            db.session.commit()
                            
                        else:
                            mission.PRIX_HT_EDL = tarif.edl_pav_villa_prix_t2
                            db.session.commit()
                            
                    if mission.TYPE_LOGEMENT[0:3] == 'PAV' and mission.TYPE_LOGEMENT[-1] == '3' :
                        tarif=Tarifs.query.filter_by(reference_client=mission.Bailleur__data.id).first()
                        if mission.CODE_FACTURATION[2:5] == '150':
                            meuble=float(tarif.edl_pav_villa_prix_t3)/2
                            mission.PRIX_HT_EDL = float(tarif.edl_pav_villa_prix_t3) + float(meuble)
                            db.session.commit()
                            
                        else:
                            mission.PRIX_HT_EDL = tarif.edl_pav_villa_prix_t3
                            db.session.commit()
                            
                
                    if mission.TYPE_LOGEMENT[0:3] == 'PAV' and mission.TYPE_LOGEMENT[-1] == '4' :
                        
                        tarif=Tarifs.query.filter_by(reference_client=mission.Bailleur__data.id).first()
                        if mission.CODE_FACTURATION[2:5] == '150':
                            meuble=float(tarif.edl_pav_villa_prix_t4)/2
                            mission.PRIX_HT_EDL = float(tarif.edl_pav_villa_prix_t4) + float(meuble)
                            db.session.commit()
                            
                        else:
                            mission.PRIX_HT_EDL = tarif.edl_pav_villa_prix_t4
                            db.session.commit()
                            
                    if mission.TYPE_LOGEMENT[0:3] == 'PAV' and mission.TYPE_LOGEMENT[-1] == '5' :
                        tarif=Tarifs.query.filter_by(reference_client=mission.Bailleur__data.id).first()
                        if mission.CODE_FACTURATION[2:5] == '150':
                            meuble=float(tarif.edl_pav_villa_prix_t5)/2
                            mission.PRIX_HT_EDL = float(tarif.edl_pav_villa_prix_t5) + float(meuble)
                            db.session.commit()
                            
                        else:
                            mission.PRIX_HT_EDL = tarif.edl_pav_villa_prix_t5
                            db.session.commit()
                            
                    if mission.TYPE_LOGEMENT[0:3] == 'PAV' and mission.TYPE_LOGEMENT[-1] == '6' :
                        tarif=Tarifs.query.filter_by(reference_client=mission.Bailleur__data.id).first()
                        if mission.CODE_FACTURATION[2:5] == '150':
                            meuble=float(tarif.edl_pav_villa_prix_t6)/2
                            mission.PRIX_HT_EDL = float(tarif.edl_pav_villa_prix_t6) + float(meuble)
                            db.session.commit()
                                
                        else:
                            mission.PRIX_HT_EDL = tarif.edl_pav_villa_prix_t6
                            db.session.commit()
                            
                    if mission.TYPE_LOGEMENT[0:3] == 'PAV' and mission.TYPE_LOGEMENT[-1] == '7' :
                        tarif=Tarifs.query.filter_by(reference_client=mission.Bailleur__data.id).first()
                        if mission.CODE_FACTURATION[2:5] == '150':
                            meuble=float(tarif.edl_pav_villa_prix_t7)/2
                            mission.PRIX_HT_EDL = float(tarif.edl_pav_villa_prix_t7) + float(meuble)
                            db.session.commit()
                            
                        else:
                            mission.PRIX_HT_EDL = tarif.edl_pav_villa_prix_t7
                            db.session.commit()
                            
                    if mission.TYPE_LOGEMENT[0:3] == 'PAV' and mission.TYPE_LOGEMENT[-1] == '8' :
                        tarif=Tarifs.query.filter_by(reference_client=mission.Bailleur__data.id).first()
                        if mission.CODE_FACTURATION[2:5] == '150':
                            meuble=float(tarif.edl_pav_villa_prix_t8)/2
                            mission.PRIX_HT_EDL = float(tarif.edl_pav_villa_prix_t8) + float(meuble)
                            db.session.commit()
                            
                        else:
                            mission.PRIX_HT_EDL = tarif.edl_pav_villa_prix_t8
                            db.session.commit()
                    else:
                       if mission.PRIX_HT_EDL == None:
                            mission.PRIX_HT_EDL = 0
                            db.session.commit()
            except:
                print("done")
                                    
            
                
            
        else:
            reference.append(vii)# make a table for historique des donnees 
            print(reference)



def failed(av):
    ba=[]
    v=0
    for oo in av:
        tr=len(oo)
    for i in range(0,tr):
        ba.append(i)
    wb = xlwt.Workbook()
    style1 = xlwt.easyxf(num_format_str='D-MMM-YY')
    ws = wb.add_sheet('Anomalie')
    for oo in av:
            for q,i in zip(ba,oo) :
                ws.write(v, q, i)
            v=v+1
    wb.save('C:/Users/user/Downloads/Telegram Desktop/Anomalie.xls')

def date_(floa,date):
        try:
            v=datetime.datetime(*xlrd.xldate_as_tuple(floa,date))
            return v
        except:
            return floa

def Missions1(loc):
    wb = xlrd.open_workbook(loc)

    sheet = wb.sheet_by_index(0)
    missions_=[]

    rows=int(sheet.nrows)
    sheet.cell_value(0, 0)
    print('ok')
    for i in range(0,rows):
        name=sheet.row_values(i)

        try:
            cli=Client.query.filter_by(reference=str(int(name[0]))).first()
        except:
            cli=None
        if cli is None:
            cli=Client.query.filter_by(nom=str(name[3]).lower()).first()
        if name[11] is None:
            SA=0
        else:
            AS=Expert.query.filter_by(nom=str(name[11].lower())).first()
            if AS is not None:
                SA= AS.id
            if AS is None :
                SA= 0
        if name[17] is None:
            IV=0
        else:
            INTERV=Expert.query.filter_by(nom=str(name[17].lower())).first()
            if INTERV is not None:
                IV= INTERV.id
            if INTERV is None  :
                IV= 0
        if name[37] is None:
            MC=0
        else:
            M_C=Expert.query.filter_by(nom=str(name[37].lower())).first()
            if M_C is not None:
                MC= M_C.id
            if M_C is None  :
                MC= 0
        if name[39] is None:
            AC=0
        else:
            A_C=Expert.query.filter_by(nom=str(name[39].lower())).first()
            if A_C is not None:
                AC= A_C.id
            if A_C is None  :
                AC= 0

        if name[69] is None:
            RCD=0
        else:
            R_CD=Expert.query.filter_by(nom=str(name[69].lower())).first()
            if R_CD is not None:
                RCD= R_CD.id
            if R_CD is None  :
                RCD= 0

        if name[71] is None:
            ACD=0
        else:
            A_CD=Expert.query.filter_by(nom=str(name[71].lower())).first()
            if A_CD is not None:
                ACD= A_CD.id
            if A_CD is None  :
                ACD= 0

        if name[73] is None:
            ACT=0
        else:
            A_CT=Expert.query.filter_by(nom=str(name[73].lower())).first()
            if A_CT is not None:
                ACT= A_CT.id
            if A_CT is None  :
                ACT= 0

        if name[75] is None:
            RCT=0
        else:
            R_CT=Expert.query.filter_by(nom=str(name[75].lower())).first()
            if R_CT is not None:
                RCT= R_CT.id
            if R_CT is None  :
                RCT= 0

        if name[77] is None:
            SCT=0
        else:
            S_CT=Expert.query.filter_by(nom=str(name[77].lower())).first()
            if S_CT is not None:
                SCT= S_CT.id
            if S_CT is None  :
                SCT= 0

        if name[79] is None:
            RCP=0
        else:
            R_CP=Expert.query.filter_by(nom=str(name[79].lower())).first()
            if R_CP is not None:
                RCP= R_CP.id
            if R_CP is None  :
                RCP= 0

        if name[81] is None:
            SCP=0
        else:
            S_CP=Expert.query.filter_by(nom=str(name[81].lower())).first()
            if S_CP is not None:
                SCP= S_CP.id
            if S_CP is None  :
                SCP= 0

        if name[83] is None:
            ASCP=0
        else:
            AS_CP=Expert.query.filter_by(nom=str(name[83].lower())).first()
            if AS_CP is not None:
                ASCP= AS_CP.id
            if AS_CP is None  :
                ASCP= 0

        if cli:
               
            mission=Mission(Reference_BAILLEUR=cli.id,
           # old=sheet["A"],
            ABONNEMENT	 = name[9] ,
            ID_AS	 = SA ,
        
          
             	
            ID_INTERV = IV ,
            NRO_FACTURE = name[8]  ,
            PRIX_HT_EDL = name[13] ,
            TVA_EDL = name[14] ,
            PRIX_TTC_EDL = name[15] ,
            Reference_LOCATAIRE	 =  name[20] ,
            Adresse1_Bien	 = name[21] ,  
            Adresse2_Bien	 = name[22] , 
            CP_Bien	 = name[23] ,  
            Ville_Bien	 = name[24] , 
            
            CA_HT_AS = name[25] , 	
            TVA_AS	 = name[26] , 
            CA_TTC_AS = name[27] , 	
            CA_HT_AC = name[28] , 	
            CA_TTC_AC	 = name[29] , 
            CA_HT_TRUST	 = name[30] , 
            TVA_TRUST	 = name[31] ,
            Prix_ht_chiffrage	 = name[33] , 
            POURCENTAGE_suiveur_chiffrage	 = name[34] ,
            POURCENTAGE_AS_chiffrage = name[35] ,	
            POURCENTAGE_manager_chiffrage  = name[36] , 	
            ID_manager_chiffrage  = MC ,
                
            POURCENTAGE_agent_chiffrage = name[38] ,	
            ID_agent_chiffrage  = AC ,	
            
            TYPE_EDL = name[40] ,	
            TITREPROPRIO = name[42] , 		
            NOMPROPRIO = name[43] , 		
            TYPE_LOGEMENT = name[27] , 	
            CODE_AMEXPERT = name[34] , 	
            COMMENTAIRE_FACTURE = name[54] , 	
            LOGEMENT_MEUBLE =name[59] , 	
            CODE_FACTURATION = name[60] , 	
            TYPE_DE_BIEN = name[61] , 	
            surface_logement1 = name[62] , 		 	
            POURCENTAGE_COM_AS_DU_CLIENT = name[68] , 	
            ID_Respon_Cell_Dev	 =RCD ,
            
            POURCENTAGE_Respon_Cell_Dev = name[70] , 	
            ID_agent_Cell_Dev = ACD, 	
            
            POURCENTAGE_Agent_Cell_Dev = name[72] ,	
            ID_Agent_CellTech = ACT,  	
            
            POURCENTAGE_Agent_Cell_Tech = name[74] , 	
            ID_Respon_Cell_Tech = RCT, #######
                
            POURCENTAGE_Respon_Cell_Tech = name[76] ,	
            ID_Suiveur_Cell_Tech  = SCT ,
            
            POURCENTAGE_Suiveur_Cell_Tech	 = name[78] , 
            ID_Respon_Cell_Planif = RCP,
            
            POURCENTAGE_Respon_Cell_Planif  = name[80] ,
            ID_Suiveur_Cell_Planif  = SCP,
            
            POURCENTAGE_Suiveur_Cell_Planif	 = name[82] , 
            ID_Agent_saisie_Cell_Planif  = ASCP,
                  
            POURCENTAGE_Agent_saisie_CEll_planif  = name[84] )
            db.session.add(mission)
            db.session.commit()
            try:
                mission.DATE_FACTURE = datetime.datetime(*xlrd.xldate_as_tuple(name[41], wb.datemode))
                db.session.commit()
            except:
                mission.DATE_FACTURE=None
                db.session.commit()
            try:
                mission.DATE_FACT_REGLEE =datetime.datetime(*xlrd.xldate_as_tuple(name[44], wb.datemode))  
                db.session.commit()
            except:
                mission.DATE_FACT_REGLEE=None
                db.session.commit()
            try:
                mission.DATE_REALISE_EDL =datetime.datetime(*xlrd.xldate_as_tuple(name[12], wb.datemode)) 
                db.session.commit()
            except:
                mission.DATE_REALISE_EDL=None
                db.session.commit()
            try:
                if mission.CODE_FACTURATION[-1]=='M':
                    print(mission.CODE_FACTURATION[2:5])
                    mission.CODE_FACTURATION[2:5] == 150
                    A=mission.CODE_FACTURATION[0:-1]
                    mission.CODE_FACTURATION = A
                    #print(mission.CODE_FACTURATION)
                    db.session.commit()

                if mission.CODE_FACTURATION[-1] == ' ':
                    print(mission.CODE_FACTURATION[2:5])
                    B=mission.CODE_FACTURATION[0:-1] 
                    mission.CODE_FACTURATION = B
                    db.session.commit()

                if mission.CODE_FACTURATION[-2:] == "00" or  mission.CODE_FACTURATION[-2:] == "50" :
                        mission.Anomalie = False
                        mission.reason = None
                        db.session.commit()
            except:
                mission.coherent = False
                mission.reason = "Anomalie bloquante codification du code facturation incorrect sur  "
               # print(mission)
                db.session.commit()
            print(mission.CODE_FACTURATION)
            print(mission.TYPE_LOGEMENT)
            try:
                if mission.TYPE_LOGEMENT[-1] == ' ':
                   # print( mission.TYPE_LOGEMENT)
                    B=mission.TYPE_LOGEMENT[0:-1] 
                    mission.TYPE_LOGEMENT = B
                    db.session.commit()


                if mission.TYPE_LOGEMENT[-1] == 'M':
                   # print(mission.TYPE_LOGEMENT)
                    mission.CODE_FACTURATION[2:5] == 150
                    #print(mission.CODE_FACTURATION)
                    B=mission.TYPE_LOGEMENT[0:-1] 
                    mission.TYPE_LOGEMENT = B
                   # print(mission.TYPE_LOGEMENT)
                    db.session.commit()

                if mission.TYPE_LOGEMENT[-2:] != mission.CODE_FACTURATION[-2:]  :
                    mission.Anomalie = True
                    mission.reason = "Anomalie non bloquante traite en  "+str(mission.TYPE_LOGEMENT[-2:])
                    db.session.commit()

                

                if len(mission.TYPE_LOGEMENT[0:4]) < 3 :
                    if len(mission.TYPE_LOGEMENT[0:2]) == 2:
                        if mission.TYPE_LOGEMENT[0] == "T":
                            B="PAV-"+str(mission.TYPE_LOGEMENT) 
                            mission.TYPE_LOGEMENT = B
                            db.session.commit()
                        if mission.TYPE_LOGEMENT[0] == "F":
                            B="APPT-"+str(mission.TYPE_LOGEMENT)
                            mission.TYPE_LOGEMENT = B
                            db.session.commit() 
                    else:
                        mission.coherent = False 
                        mission.reason = "Anomalie bloquante codification du type de logement incorrect sur  "+str(mission.TYPE_LOGEMENT)
                        db.session.commit()
                
            except:
                mission.coherent = False
                mission.reason = "Anomalie bloquante codification du type de logement incorrect "
                db.session.commit()
            try:
                if mission.TYPE_LOGEMENT[-1] == ' ':
                   # print( mission.TYPE_LOGEMENT)
                    B=mission.TYPE_LOGEMENT[0:-1] 
                    mission.TYPE_LOGEMENT = B
                    db.session.commit()


                if mission.TYPE_LOGEMENT[-1] == 'M':
                   # print(mission.TYPE_LOGEMENT)
                    mission.CODE_FACTURATION[2:5] == 150
                    #print(mission.CODE_FACTURATION)
                    B=mission.TYPE_LOGEMENT[0:-1] 
                    mission.TYPE_LOGEMENT = B
                   # print(mission.TYPE_LOGEMENT)
                    db.session.commit()

                if mission.TYPE_LOGEMENT[-2:] != mission.CODE_FACTURATION[-2:]  :
                    mission.Anomalie = True
                    mission.reason = "Anomalie non bloquante traite en  "+str(mission.TYPE_LOGEMENT[-2:])
                    db.session.commit()

                

                if len(mission.TYPE_LOGEMENT[0:4]) < 3 :
                    if len(mission.TYPE_LOGEMENT[0:2]) == 2:
                        if mission.TYPE_LOGEMENT[0] == "T":
                            B="PAV-"+str(mission.TYPE_LOGEMENT) 
                            mission.TYPE_LOGEMENT = B
                            db.session.commit()
                        if mission.TYPE_LOGEMENT[0] == "F":
                            B="APPT-"+str(mission.TYPE_LOGEMENT)
                            mission.TYPE_LOGEMENT = B
                            db.session.commit() 
                    else:
                        mission.coherent = False 
                        mission.reason = "Anomalie bloquante codification du type de logement incorrect sur  "+str(mission.TYPE_LOGEMENT)
                        db.session.commit()
                
            except:
                mission.coherent = False
                mission.reason = "Anomalie bloquante codification du type de logement incorrect "
                db.session.commit()
            
        else:

            missions_.append([name[0],name[1],name[2],name[3],name[4],name[5],name[6],name[7],
            name[8],name[9],name[10],name[11],date_(name[12],wb.datemode),name[13]
            ,name[14],name[15],name[16],
            name[17],name[18],name[19],name[20],name[21],name[22],name[23],name[24],
            name[25],name[26],name[27],name[28],name[29],name[30],name[31],name[32],
            name[33],name[34],name[35],name[36],name[37],name[38],name[39],name[40],
                date_(name[41],wb.datemode),name[42],
            name[43],date_(name[44],wb.datemode),
            name[45],name[46],name[47],name[48],
                name[49],name[50],name[51],name[52],name[53],name[54],name[55],name[56],
                name[57],name[58],name[59],name[60],name[61],name[62],name[63],name[64],
                name[65],name[66],name[67],name[68],name[69],name[70],name[71],name[72],
                name[73],name[74],name[75],name[76],name[77],name[78],name[79],name[80],
                    name[81],name[82],name[83],name[84]])
    print(missions_)
    failed(missions_)
    






















#Missions()






#insert_client('Bailleur')
#insert_client('Locataire')
#insert_client('Prop')
#expert_('Interv')
#expert_('CONCESS')
#expert_('Manager_chiffrage')
#expert_('Agent_chiffrage')
#expert_('Respon Cell Dev')
#expert_('agent Cell Dev')
#expert_('Agent CellTech')
#expert_('Respon Cell Tech')
#expert_('Suiveur Cell Tech')
#expert_('Respon Cell Planif')
#expert_('Suiveur Cell Planif')
#expert_('Agent saisie Cell Planif')
