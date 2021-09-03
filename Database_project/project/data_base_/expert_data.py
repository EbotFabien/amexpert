import sys
import datetime
import openpyxl
import xlrd,xlwt
from Database_project.project.data_base_ import db
from Database_project.project.data_base_.Models import Tarifs,Mission,Client,Expert,Client_History,prospect,prospect_History,Expert_History,Tarif_base
import pandas as pd
from Database_project.project.data_base_.client_data  import regex1

def failed1(av):
    ba=[]
    v=0
    for oo in av:
        tr=len(oo)
    for i in range(0,tr):
        ba.append(i)
    wb = xlwt.Workbook()
    style1 = xlwt.easyxf(num_format_str='D-MMM-YY')
    ws = wb.add_sheet('Anomalie')
    for oo in av:
            for q,i in zip(ba,oo) :
                if isinstance(i,datetime.datetime) == True:
                    ws.write(v, q, i,style1)
                elif isinstance(i,datetime.date) == True:
                    ws.write(v, q, i,style1) 
                else:
                    ws.write(v, q, i)
            v=v+1
    wb.save('C:/Users/user/Downloads/Telegram Desktop/Anomalieexpert.xls')

def xpert(loc):
    wb_obj = openpyxl.load_workbook(loc,data_only=True)
    anomalie=[]
    sheet=wb_obj.active
    for i in range(0,sheet.max_row):
        n=sheet["M"][i].value
        if type(n) == str:
            cli=Expert.query.filter_by(nom=str(sheet["M"][i].value.lower())).first()
            if cli  is None:
                exp=Expert()
                his=Expert_History()
                
                exp.nom=regex1(sheet["M"][i].value,'str1')
                exp.numero=regex1(sheet["S"][i].value,'str1')
                exp.TYPE=regex1(sheet["I"][i].value,'str1')
                exp.email=regex1(sheet["T"][i].value,'str1')
                exp.email_perso=regex1(sheet["U"][i].value,'str1')
                exp.code_tva=regex1(sheet["Q"][i].value,'str1')
                taux_tva=regex1(sheet["R"][i].value,'int')
                if taux_tva == False:
                    reason="Erreur sur Siret"
                    anomalie.append([sheet["A"][i].value,sheet["B"][i].value,sheet["C"][i].value,
                                    sheet["D"][i].value,sheet["E"][i].value,sheet["F"][i].value,sheet["G"][i].value,
                                    sheet["H"][i].value,sheet["I"][i].value,sheet["J"][i].value,sheet["K"][i].value,
                                    sheet["L"][i].value,sheet["M"][i].value,sheet["N"][i].value,sheet["O"][i].value,
                                    sheet["P"][i].value,sheet["Q"][i].value,sheet["R"][i].value,sheet["S"][i].value,
                                    sheet["T"][i].value,sheet["U"][i].value,sheet["V"][i].value,sheet["W"][i].value,
                                    sheet["X"][i].value,sheet["Y"][i].value,sheet["Z"][i].value,sheet["AA"][i].value,reason)
                else:
                    exp.taux_tva=taux_tva
                siret=regex1(sheet["L"][i].value,'int')
                if siret == False:
                    reason="Erreur sur Siret"
                    anomalie.append([sheet["A"][i].value,sheet["B"][i].value,sheet["C"][i].value,
                                    sheet["D"][i].value,sheet["E"][i].value,sheet["F"][i].value,sheet["G"][i].value,
                                    sheet["H"][i].value,sheet["I"][i].value,sheet["J"][i].value,sheet["K"][i].value,
                                    sheet["L"][i].value,sheet["M"][i].value,sheet["N"][i].value,sheet["O"][i].value,
                                    sheet["P"][i].value,sheet["Q"][i].value,sheet["R"][i].value,sheet["S"][i].value,
                                    sheet["T"][i].value,sheet["U"][i].value,sheet["V"][i].value,sheet["W"][i].value,
                                    sheet["X"][i].value,sheet["Y"][i].value,sheet["Z"][i].value,sheet["AA"][i].value,reason)
                else:
                    exp.siret=siret
                date_entree=regex1(sheet["D"][i].value,'date')
                if date_entree == False:
                    reason="Erreur sur date entree"
                    anomalie.append([sheet["A"][i].value,sheet["B"][i].value,sheet["C"][i].value,
                                    sheet["D"][i].value,sheet["E"][i].value,sheet["F"][i].value,sheet["G"][i].value,
                                    sheet["H"][i].value,sheet["I"][i].value,sheet["J"][i].value,sheet["K"][i].value,
                                    sheet["L"][i].value,sheet["M"][i].value,sheet["N"][i].value,sheet["O"][i].value,
                                    sheet["P"][i].value,sheet["Q"][i].value,sheet["R"][i].value,sheet["S"][i].value,
                                    sheet["T"][i].value,sheet["U"][i].value,sheet["V"][i].value,sheet["W"][i].value,
                                    sheet["X"][i].value,sheet["Y"][i].value,sheet["Z"][i].value,sheet["AA"][i].value,reason)
                else:
                    exp.date_entree=date_entree
                exp.trigramme=regex1(sheet["M"][i].value,'str1')    
                
                his.secteur=regex1(sheet["J"][i].value,'str1') 
                his.adresse1=regex1(sheet["M"][i].value,'str1') 
                his.adresse2=regex1(sheet["N"][i].value,'str1') 
                cp=regex1(sheet["O"][i].value,'int') 
                if cp==False:
                    reason="Erreur sur CP"
                    anomalie.append([sheet["A"][i].value,sheet["B"][i].value,sheet["C"][i].value,
                                    sheet["D"][i].value,sheet["E"][i].value,sheet["F"][i].value,sheet["G"][i].value,
                                    sheet["H"][i].value,sheet["I"][i].value,sheet["J"][i].value,sheet["K"][i].value,
                                    sheet["L"][i].value,sheet["M"][i].value,sheet["N"][i].value,sheet["O"][i].value,
                                    sheet["P"][i].value,sheet["Q"][i].value,sheet["R"][i].value,sheet["S"][i].value,
                                    sheet["T"][i].value,sheet["U"][i].value,sheet["V"][i].value,sheet["W"][i].value,
                                    sheet["X"][i].value,sheet["Y"][i].value,sheet["Z"][i].value,sheet["AA"][i].value,reason)
                else:
                    his.cp=cp
                his.ville=regex1(sheet["P"][i].value,'str1') 
                his.login_backof=regex1(sheet["P"][i].value,'str1') 
                his.pwd_backof=regex1(sheet["P"][i].value,'str1') 
                his.login_extranet=regex1(sheet["P"][i].value,'str1') 
                his.pwd_extranet=regex1(sheet["P"][i].value,'str1') 
                his.login_google=regex1(sheet["P"][i].value,'str1') 
                his.pwd_google=regex1(sheet["P"][i].value,'str1') 
                his.observations_de_suivi=regex1(sheet["P"][i].value,'str1') 
                his.actif_parti=regex1(sheet["P"][i].value,'str1') 
                his.type_certification=regex1(sheet["P"][i].value,'str1') 
                his.date_certification_initiale=regex1(sheet["P"][i].value,'str1') 
                his.date_renouv_certification=regex1(sheet["P"][i].value,'str1') 
                his.pwd_gsuite=regex1(sheet["P"][i].value,'str1') 
                db.session.add(exp)
                db.session.add(his)
                db.session.commit()
                expert_id=expert.id
                db.session.commit()
                
    failed1(anomalie)
    
